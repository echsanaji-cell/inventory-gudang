import requests
import os
import barcode
import smtplib
import xlrd
import re
import secrets as secrets_lib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from barcode.writer import ImageWriter
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from openpyxl import load_workbook
import io
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, session, flash
from werkzeug.security import check_password_hash, generate_password_hash
from functools import wraps
from db import get_db_connection
from reportlab.lib.styles import ParagraphStyle
from google.oauth2 import service_account
from googleapiclient.discovery import build
import json as json_lib

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-key")
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('RENDER') is not None
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=60)
from flask_wtf import CSRFProtect
from datetime import timedelta

csrf = CSRFProtect(app)

def baca_excel_universal(file_storage):
    """Baca file Excel .xls maupun .xlsx, kembalikan list of rows (list of list)."""
    nama_file = file_storage.filename.lower()

    if nama_file.endswith('.xls'):
        isi = file_storage.read()
        wb = xlrd.open_workbook(file_contents=isi)
        sheet = wb.sheet_by_index(0)
        rows = []
        for r in range(sheet.nrows):
            rows.append([sheet.cell_value(r, c) if sheet.cell_value(r, c) != '' else None for c in range(sheet.ncols)])
        return rows
    else:
        wb = load_workbook(file_storage, data_only=True)
        ws = wb.worksheets[0]
        return [list(row) for row in ws.iter_rows(values_only=True)]

def kirim_email_stok_kritis():
    mail_username = os.environ.get('MAIL_USERNAME')
    mail_password = os.environ.get('MAIL_PASSWORD')
    mail_to = os.environ.get('MAIL_TO')

    if not all([mail_username, mail_password, mail_to]):
        return False, "Konfigurasi email belum lengkap di environment variables."

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """SELECT * FROM buku 
           WHERE stok_minimum > 0 AND stok <= stok_minimum 
           ORDER BY stok ASC"""
    )
    stok_menipis = cur.fetchall()
    cur.close()
    conn.close()

    if not stok_menipis:
        return True, "Tidak ada buku dengan stok menipis, email tidak dikirim."

    baris_tabel = ""
    for buku in stok_menipis:
        baris_tabel += f"""
        <tr>
            <td style="padding:6px 10px; border-bottom:1px solid #ddd;">{buku['judul']}</td>
            <td style="padding:6px 10px; border-bottom:1px solid #ddd;">{buku['isbn']}</td>
            <td style="padding:6px 10px; border-bottom:1px solid #ddd; color:#A63A2C; font-weight:bold;">{buku['stok']}</td>
            <td style="padding:6px 10px; border-bottom:1px solid #ddd;">{buku['stok_minimum']}</td>
        </tr>"""

    html_body = f"""
    <html><body style="font-family: Arial, sans-serif;">
        <h2>⚠️ Peringatan Stok Menipis — Inventory Gudang</h2>
        <p>Ada {len(stok_menipis)} buku dengan stok di bawah atau sama dengan batas minimum:</p>
        <table style="border-collapse: collapse; width:100%;">
            <tr style="background:#202A33; color:white;">
                <th style="padding:6px 10px; text-align:left;">Judul</th>
                <th style="padding:6px 10px; text-align:left;">ISBN</th>
                <th style="padding:6px 10px; text-align:left;">Stok</th>
                <th style="padding:6px 10px; text-align:left;">Minimum</th>
            </tr>
            {baris_tabel}
        </table>
        <p style="margin-top:20px; color:#888; font-size:12px;">
            Email otomatis dari sistem Inventory Gudang — dikirim {datetime.now().strftime('%d-%m-%Y %H:%M')}
        </p>
    </body></html>
    """

    msg = MIMEMultipart('alternative')
    msg['Subject'] = f'⚠️ {len(stok_menipis)} Buku Stok Menipis - Inventory Gudang'
    msg['From'] = mail_username
    msg['To'] = mail_to
    msg.attach(MIMEText(html_body, 'html'))

    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(mail_username, mail_password)
        server.sendmail(mail_username, mail_to.split(','), msg.as_string())
        server.quit()
        return True, f"Email berhasil dikirim ke {mail_to} ({len(stok_menipis)} buku)."
    except Exception as e:
        return False, f"Gagal kirim email: {str(e)}"

def ambil_semua_data_backup():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM buku ORDER BY id ASC")
    data_buku = cur.fetchall()

    cur.execute("SELECT * FROM transaksi ORDER BY id ASC")
    data_transaksi = cur.fetchall()

    cur.execute("SELECT * FROM users ORDER BY id ASC")
    data_users = cur.fetchall()

    cur.close()
    conn.close()
    return data_buku, data_transaksi, data_users


def buat_backup_excel():
    data_buku, data_transaksi, data_users = ambil_semua_data_backup()

    wb = Workbook()

    ws_buku = wb.active
    ws_buku.title = "Buku"
    if data_buku:
        headers = list(data_buku[0].keys())
        ws_buku.append(headers)
        for row in data_buku:
            ws_buku.append([str(row[h]) if row[h] is not None else '' for h in headers])

    ws_transaksi = wb.create_sheet("Transaksi")
    if data_transaksi:
        headers = list(data_transaksi[0].keys())
        ws_transaksi.append(headers)
        for row in data_transaksi:
            ws_transaksi.append([str(row[h]) if row[h] is not None else '' for h in headers])

    ws_users = wb.create_sheet("Users")
    if data_users:
        headers = [h for h in data_users[0].keys() if h != 'password_hash']
        ws_users.append(headers)
        for row in data_users:
            ws_users.append([str(row[h]) if row[h] is not None else '' for h in headers])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def buat_backup_json():
    data_buku, data_transaksi, data_users = ambil_semua_data_backup()

    def convert_row(row):
        hasil = {}
        for k, v in dict(row).items():
            if hasattr(v, 'isoformat'):
                hasil[k] = v.isoformat()
            else:
                hasil[k] = v
        return hasil

    backup = {
        'dibuat_pada': datetime.now().isoformat(),
        'buku': [convert_row(r) for r in data_buku],
        'transaksi': [convert_row(r) for r in data_transaksi],
        'users': [convert_row(r) for r in data_users],
    }

    output = io.BytesIO()
    output.write(json_lib.dumps(backup, indent=2, ensure_ascii=False).encode('utf-8'))
    output.seek(0)
    return output

def restore_dari_backup(data_backup):
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        # hapus data lama, urutan penting karena foreign key
        cur.execute("DELETE FROM transaksi")
        cur.execute("DELETE FROM buku")
        cur.execute("DELETE FROM users")

        # restore users
        for u in data_backup.get('users', []):
            cur.execute(
                """INSERT INTO users (id, username, password_hash, nama_lengkap, role, is_active, failed_attempts, locked_until, created_at)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (u['id'], u['username'], u['password_hash'], u.get('nama_lengkap'),
                 u.get('role', 'staff'), u.get('is_active', True), u.get('failed_attempts', 0),
                 u.get('locked_until'), u.get('created_at'))
            )

        # restore buku
        for b in data_backup.get('buku', []):
            cur.execute(
                """INSERT INTO buku (id, isbn, judul, penulis, penerbit, sampul_url, stok, stok_minimum,
                                      jumlah_rencana, tanggal_masuk, catatan, created_at, updated_at)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (b['id'], b['isbn'], b['judul'], b.get('penulis'), b.get('penerbit'), b.get('sampul_url'),
                 b.get('stok', 0), b.get('stok_minimum', 0), b.get('jumlah_rencana', 0),
                 b.get('tanggal_masuk'), b.get('catatan'), b.get('created_at'), b.get('updated_at'))
            )

        # restore transaksi
        for t in data_backup.get('transaksi', []):
            cur.execute(
                """INSERT INTO transaksi (id, buku_id, tipe, jumlah, keterangan, pihak_terkait, user_id, tanggal,
                                           jenis_keluar, tanggal_kembali_rencana, tanggal_kembali_aktual, created_at)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (t['id'], t['buku_id'], t['tipe'], t['jumlah'], t.get('keterangan'), t.get('pihak_terkait'),
                 t.get('user_id'), t.get('tanggal'), t.get('jenis_keluar'),
                 t.get('tanggal_kembali_rencana'), t.get('tanggal_kembali_aktual'), t.get('created_at'))
            )

        # reset sequence ID biar data baru berikutnya nggak bentrok sama ID hasil restore
        for tabel in ['buku', 'transaksi', 'users']:
            cur.execute(
                f"SELECT setval(pg_get_serial_sequence('{tabel}', 'id'), COALESCE((SELECT MAX(id) FROM {tabel}), 1))"
            )

        conn.commit()
        jumlah = f"{len(data_backup.get('buku', []))} buku, {len(data_backup.get('transaksi', []))} transaksi, {len(data_backup.get('users', []))} users"
        return True, f"Restore berhasil: {jumlah}."
    except Exception as e:
        conn.rollback()
        return False, f"Restore GAGAL, data dikembalikan ke kondisi semula: {str(e)}"
    finally:
        cur.close()
        conn.close()

def kirim_backup_email():
    mail_username = os.environ.get('MAIL_USERNAME')
    mail_password = os.environ.get('MAIL_PASSWORD')
    mail_to = os.environ.get('MAIL_TO')

    if not all([mail_username, mail_password, mail_to]):
        return False, "Konfigurasi email belum lengkap di environment variables."

    try:
        backup_file = buat_backup_json()
        nama_file = f"backup-inventory-gudang-{datetime.now().strftime('%Y%m%d-%H%M')}.json"

        msg = MIMEMultipart()
        msg['Subject'] = f'📦 Backup Otomatis Inventory Gudang - {datetime.now().strftime("%d-%m-%Y")}'
        msg['From'] = mail_username
        msg['To'] = mail_to
        msg.attach(MIMEText(
            f"Backup otomatis dibuat pada {datetime.now().strftime('%d-%m-%Y %H:%M')}.\n\n"
            f"File terlampir berisi seluruh data buku, transaksi, dan users (password terenkripsi).",
            'plain'
        ))

        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(backup_file.read())
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', f'attachment; filename="{nama_file}"')
        msg.attach(attachment)

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(mail_username, mail_password)
        server.sendmail(mail_username, mail_to.split(','), msg.as_string())
        server.quit()

        return True, f"Backup berhasil dikirim ke {mail_to}."
    except Exception as e:
        return False, f"Gagal membuat/mengirim backup: {str(e)}"

def ambil_int(form, nama_field, default=0):
    """Ambil nilai integer dari form dengan aman, fallback ke default kalau kosong/tidak valid"""
    nilai = form.get(nama_field, '').strip()
    if not nilai:
        return default
    try:
        return int(float(nilai))
    except (ValueError, TypeError):
        return default
    
def catat_aktivitas(aksi, detail='', buku_id=None):
    """Catat aktivitas penting ke activity_log. Gagal diam-diam kalau error, tidak boleh ganggu proses utama."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO activity_log (user_id, username, aksi, detail, buku_id) VALUES (%s, %s, %s, %s, %s)",
            (session.get('user_id'), session.get('username'), aksi, detail, buku_id)
        )
        conn.commit()
        cur.close()
        conn.close()
    except Exception:
        pass  # audit log tidak boleh bikin fitur utama gagal
    
def sync_ke_google_sheets():
    creds_json = os.environ.get('GOOGLE_SHEETS_CREDENTIALS')
    sheet_id = os.environ.get('GOOGLE_SHEETS_ID')

    if not creds_json or not sheet_id:
        return False, "Konfigurasi Google Sheets belum lengkap di environment variables."

    try:
        creds_dict = json_lib.loads(creds_json)
        creds = service_account.Credentials.from_service_account_info(
            creds_dict, scopes=['https://www.googleapis.com/auth/spreadsheets']
        )
        service = build('sheets', 'v4', credentials=creds)
    except Exception as e:
        return False, f"Gagal autentikasi Google Sheets: {str(e)}"

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM buku ORDER BY judul ASC")
    daftar_buku = cur.fetchall()
    cur.close()
    conn.close()

    header = ['No', 'ISBN', 'Judul', 'Penulis', 'Penerbit', 'Diterima', 'Rencana', 'Status', 'Stok Minimum', 'Tanggal Masuk', 'Catatan']

    def hitung_status(buku):
        if buku['jumlah_rencana'] <= 0:
            return 'Tanpa Rencana'
        if buku['stok'] == 0:
            return 'Belum Ada'
        elif buku['stok'] < buku['jumlah_rencana']:
            return 'Sebagian'
        elif buku['stok'] == buku['jumlah_rencana']:
            return 'Lengkap'
        else:
            return 'Lebih'

    def buat_rows(buku_list):
        rows = [header]
        for i, buku in enumerate(buku_list, start=1):
            rows.append([
                i, buku['isbn'], buku['judul'], buku['penulis'] or '-',
                buku['penerbit'] or '-', buku['stok'], buku['jumlah_rencana'],
                hitung_status(buku), buku['stok_minimum'],
                str(buku['tanggal_masuk']) if buku['tanggal_masuk'] else '-',
                buku['catatan'] or '-'
            ])
        return rows

    def sanitize_nama_tab(nama):
        nama = nama or 'Tanpa Penerbit'
        for ch in [':', '\\', '/', '?', '*', '[', ']']:
            nama = nama.replace(ch, '-')
        return nama[:95]

    # filter buku dengan status "Sebagian", urutkan per-penerbit lalu alfabet judul
    daftar_sebagian = sorted(
        [b for b in daftar_buku if b['jumlah_rencana'] > 0 and 0 < b['stok'] < b['jumlah_rencana']],
        key=lambda b: ((b['penerbit'] or '').lower(), b['judul'].lower())
    )

    # filter buku dengan status "Lebih", urutkan per-penerbit lalu alfabet judul
    daftar_lebih = sorted(
        [b for b in daftar_buku if b['jumlah_rencana'] > 0 and b['stok'] > b['jumlah_rencana']],
        key=lambda b: ((b['penerbit'] or '').lower(), b['judul'].lower())
    )
    # rekap per Tanggal Masuk + Penerbit, HANYA yang status Sebagian/Lengkap (match dashboard)
    dari_rekap = {}
    for b in daftar_buku:
        if b['jumlah_rencana'] > 0 and b['stok'] > 0 and b['stok'] <= b['jumlah_rencana']:
            tgl = str(b['tanggal_masuk']) if b['tanggal_masuk'] else '(Tanpa Tanggal)'
            penerbit = b['penerbit'] or '(Tanpa Penerbit)'
            key = (tgl, penerbit)
            dari_rekap[key] = dari_rekap.get(key, 0) + 1

    rekap_tanggal_rows = [['Tanggal Masuk', 'Penerbit', 'Jumlah Judul']]
    for (tgl, penerbit) in sorted(dari_rekap.keys(), key=lambda k: (k[0], k[1].lower())):
        rekap_tanggal_rows.append([tgl, penerbit, dari_rekap[(tgl, penerbit)]])

    total_rekap = sum(dari_rekap.values())
    rekap_tanggal_rows.append(['TOTAL', '', total_rekap])
    # hitung ringkasan status per penerbit (persis logic Dashboard)
    ringkasan_per_penerbit = {}
    for b in daftar_buku:
        nama_p = b['penerbit'] or '(Tanpa Penerbit)'
        if nama_p not in ringkasan_per_penerbit:
            ringkasan_per_penerbit[nama_p] = {'belum_ada': 0, 'sebagian': 0, 'lengkap': 0, 'lebih': 0, 'tanpa_rencana': 0}

        if b['jumlah_rencana'] <= 0:
            ringkasan_per_penerbit[nama_p]['tanpa_rencana'] += 1
        elif b['stok'] == 0:
            ringkasan_per_penerbit[nama_p]['belum_ada'] += 1
        elif b['stok'] < b['jumlah_rencana']:
            ringkasan_per_penerbit[nama_p]['sebagian'] += 1
        elif b['stok'] == b['jumlah_rencana']:
            ringkasan_per_penerbit[nama_p]['lengkap'] += 1
        else:
            ringkasan_per_penerbit[nama_p]['lebih'] += 1

    ringkasan_header = ['Penerbit', 'Belum Ada', 'Sebagian', 'Lengkap', 'Lebih', 'Tanpa Rencana', 'Total Judul']
    ringkasan_rows = [ringkasan_header]
    total_keseluruhan = {'belum_ada': 0, 'sebagian': 0, 'lengkap': 0, 'lebih': 0, 'tanpa_rencana': 0}

    for nama_p in sorted(ringkasan_per_penerbit.keys()):
        s = ringkasan_per_penerbit[nama_p]
        total_judul_p = s['belum_ada'] + s['sebagian'] + s['lengkap'] + s['lebih'] + s['tanpa_rencana']
        ringkasan_rows.append([nama_p, s['belum_ada'], s['sebagian'], s['lengkap'], s['lebih'], s['tanpa_rencana'], total_judul_p])
        for k in total_keseluruhan:
            total_keseluruhan[k] += s[k]

    total_semua = sum(total_keseluruhan.values())
    ringkasan_rows.append([
        'TOTAL KESELURUHAN', total_keseluruhan['belum_ada'], total_keseluruhan['sebagian'],
        total_keseluruhan['lengkap'], total_keseluruhan['lebih'], total_keseluruhan['tanpa_rencana'], total_semua
    ])
    # data distribusi ke tujuan (rencana vs terkirim, per tujuan per buku)
    conn2 = get_db_connection()
    cur2 = conn2.cursor()
    cur2.execute(
        """SELECT t.nama as nama_tujuan, t.kecamatan, t.kabupaten_kota,
                  b.isbn, b.judul, b.penerbit, dr.jumlah_rencana,
                  COALESCE((
                      SELECT SUM(tr.jumlah) FROM transaksi tr
                      WHERE tr.tujuan_id = dr.tujuan_id AND tr.buku_id = dr.buku_id AND tr.tipe = 'keluar'
                  ), 0) as jumlah_terkirim
           FROM distribusi_rencana dr
           JOIN tujuan t ON dr.tujuan_id = t.id
           JOIN buku b ON dr.buku_id = b.id
           ORDER BY t.nama ASC, b.judul ASC"""
    )
    data_distribusi = cur2.fetchall()

    cur2.execute(
        """SELECT t.nama, t.kecamatan, t.kabupaten_kota,
                  COALESCE(SUM(dr.jumlah_rencana), 0) as rencana,
                  COALESCE((
                      SELECT SUM(tr.jumlah) FROM transaksi tr WHERE tr.tujuan_id = t.id AND tr.tipe = 'keluar'
                  ), 0) as terkirim
           FROM tujuan t
           LEFT JOIN distribusi_rencana dr ON dr.tujuan_id = t.id
           GROUP BY t.id, t.nama, t.kecamatan, t.kabupaten_kota
           ORDER BY t.nama ASC"""
    )
    data_ringkasan_tujuan = cur2.fetchall()
    cur2.close()
    conn2.close()

    distribusi_header = ['Tujuan', 'Kecamatan', 'Kabupaten/Kota', 'ISBN', 'Judul', 'Penerbit', 'Rencana', 'Terkirim', 'Status']
    distribusi_rows = [distribusi_header]
    for d in data_distribusi:
        if d['jumlah_terkirim'] >= d['jumlah_rencana']:
            status = 'Lengkap'
        elif d['jumlah_terkirim'] > 0:
            status = 'Sebagian'
        else:
            status = 'Belum'
        distribusi_rows.append([
            d['nama_tujuan'], d['kecamatan'] or '-', d['kabupaten_kota'] or '-',
            d['isbn'], d['judul'], d['penerbit'] or '-', d['jumlah_rencana'], d['jumlah_terkirim'], status
        ])

    ringkasan_tujuan_header = ['Tujuan', 'Kecamatan', 'Kabupaten/Kota', 'Rencana', 'Terkirim', 'Persen', 'Status']
    ringkasan_tujuan_rows = [ringkasan_tujuan_header]
    for t in data_ringkasan_tujuan:
        persen = round((t['terkirim'] / t['rencana'] * 100), 1) if t['rencana'] > 0 else 0
        if t['rencana'] == 0:
            status_t = 'Belum Ada Rencana'
        elif t['terkirim'] == 0:
            status_t = 'Belum Dikirim'
        elif t['terkirim'] < t['rencana']:
            status_t = 'Sebagian'
        else:
            status_t = 'Lengkap'
        ringkasan_tujuan_rows.append([
            t['nama'], t['kecamatan'] or '-', t['kabupaten_kota'] or '-',
            t['rencana'], t['terkirim'], f"{persen}%", status_t
        ])
    try:
        sheet = service.spreadsheets()

        meta = sheet.get(spreadsheetId=sheet_id).execute()
        tab_ada = [s['properties']['title'] for s in meta['sheets']]

        tab_baru_dibutuhkan = []
        if 'Sebagian' not in tab_ada:
            tab_baru_dibutuhkan.append({'addSheet': {'properties': {'title': 'Sebagian'}}})
        if 'Lebih' not in tab_ada:
            tab_baru_dibutuhkan.append({'addSheet': {'properties': {'title': 'Lebih'}}})
        if 'Ringkasan Status' not in tab_ada:
            tab_baru_dibutuhkan.append({'addSheet': {'properties': {'title': 'Ringkasan Status'}}})
        if 'Rekap per Tanggal' not in tab_ada:
            tab_baru_dibutuhkan.append({'addSheet': {'properties': {'title': 'Rekap per Tanggal'}}})
        if 'Distribusi Tujuan' not in tab_ada:
            tab_baru_dibutuhkan.append({'addSheet': {'properties': {'title': 'Distribusi Tujuan'}}})
        if 'Ringkasan Tujuan' not in tab_ada:
            tab_baru_dibutuhkan.append({'addSheet': {'properties': {'title': 'Ringkasan Tujuan'}}})

        if tab_baru_dibutuhkan:
            sheet.batchUpdate(
                spreadsheetId=sheet_id, body={'requests': tab_baru_dibutuhkan}
            ).execute()

        sheet.values().batchClear(
            spreadsheetId=sheet_id,
            body={'ranges': ['Sheet1', "'Sebagian'", "'Lebih'", "'Ringkasan Status'", "'Rekap per Tanggal'", "'Distribusi Tujuan'", "'Ringkasan Tujuan'"]}
        ).execute()

        sheet.values().batchUpdate(
            spreadsheetId=sheet_id,
            body={
                'valueInputOption': 'RAW',
                'data': [
                    {'range': 'Sheet1!A1', 'values': buat_rows(daftar_buku)},
                    {'range': "'Sebagian'!A1", 'values': buat_rows(daftar_sebagian)},
                    {'range': "'Lebih'!A1", 'values': buat_rows(daftar_lebih)},
                    {'range': "'Ringkasan Status'!A1", 'values': ringkasan_rows},
                    {'range': "'Rekap per Tanggal'!A1", 'values': rekap_tanggal_rows},
                    {'range': "'Distribusi Tujuan'!A1", 'values': distribusi_rows},
                    {'range': "'Ringkasan Tujuan'!A1", 'values': ringkasan_tujuan_rows}
                ]
            }
        ).execute()

        return True, f"Berhasil sync {len(daftar_buku)} buku dan {len(data_distribusi)} baris data distribusi ke {len(tab_baru_dibutuhkan) + len(tab_ada)} tab."
    except Exception as e:
        return False, f"Gagal menulis ke Google Sheets: {str(e)}"
# ------------------ DECORATOR: wajib login ------------------
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        session.permanent = True
        session.modified = True  # refresh waktu expired tiap ada request, biar user aktif nggak ke-logout
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Halaman ini khusus untuk admin.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function
def viewer_blocked(f):
    """Blokir role viewer dari akses fitur selain Dashboard & Data Buku"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('role') == 'viewer':
            flash('Akun ini hanya memiliki akses lihat Data Buku.', 'danger')
            return redirect(url_for('buku_list'))
        return f(*args, **kwargs)
    return decorated_function
# ------------------ LOGIN ------------------
MAX_FAILED_ATTEMPTS = 5
LOCKOUT_MINUTES = 15

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT * FROM users WHERE username = %s AND is_active = TRUE",
            (username,)
        )
        user = cur.fetchone()

        if not user:
            flash('Username atau password salah.', 'danger')
            cur.close()
            conn.close()
            return render_template('login.html')

        # cek apakah akun sedang terkunci
        if user['locked_until'] and user['locked_until'] > datetime.now():
            sisa_menit = int((user['locked_until'] - datetime.now()).total_seconds() / 60) + 1
            flash(f'Akun terkunci karena terlalu banyak percobaan gagal. Coba lagi dalam {sisa_menit} menit.', 'danger')
            cur.close()
            conn.close()
            return render_template('login.html')

        if check_password_hash(user['password_hash'], password):
            # login berhasil: reset failed_attempts
            cur.execute(
                "UPDATE users SET failed_attempts = 0, locked_until = NULL WHERE id = %s",
                (user['id'],)
            )
            conn.commit()

            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            session['nama_lengkap'] = user['nama_lengkap']

            cur.close()
            conn.close()
            return redirect(url_for('dashboard'))
        else:
            # login gagal: tambah counter
            new_attempts = user['failed_attempts'] + 1

            if new_attempts >= MAX_FAILED_ATTEMPTS:
                locked_until = datetime.now() + timedelta(minutes=LOCKOUT_MINUTES)
                cur.execute(
                    "UPDATE users SET failed_attempts = %s, locked_until = %s WHERE id = %s",
                    (new_attempts, locked_until, user['id'])
                )
                conn.commit()
                flash(f'Terlalu banyak percobaan gagal. Akun dikunci selama {LOCKOUT_MINUTES} menit.', 'danger')
            else:
                cur.execute(
                    "UPDATE users SET failed_attempts = %s WHERE id = %s",
                    (new_attempts, user['id'])
                )
                conn.commit()
                sisa = MAX_FAILED_ATTEMPTS - new_attempts
                flash(f'Username atau password salah. Percobaan tersisa: {sisa}.', 'danger')

        cur.close()
        conn.close()

    return render_template('login.html')

# ------------------ LOGOUT ------------------
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ------------------ DASHBOARD ------------------
@app.route('/')
@login_required
def dashboard():
    conn = get_db_connection()
    cur = conn.cursor()

    # total buku, total stok, total rencana
    cur.execute(
        """SELECT COUNT(*) as total_judul, 
                  COALESCE(SUM(stok), 0) as total_stok,
                  COALESCE(SUM(jumlah_rencana), 0) as total_rencana
           FROM buku"""
    )
    ringkasan = cur.fetchone()

    # hitung jumlah buku per status penerimaan
    cur.execute(
        """SELECT 
             COUNT(*) FILTER (WHERE jumlah_rencana > 0 AND stok = 0) as belum_ada,
             COUNT(*) FILTER (WHERE jumlah_rencana > 0 AND stok > 0 AND stok < jumlah_rencana) as sebagian,
             COUNT(*) FILTER (WHERE jumlah_rencana > 0 AND stok = jumlah_rencana) as lengkap,
             COUNT(*) FILTER (WHERE jumlah_rencana > 0 AND stok > jumlah_rencana) as lebih
           FROM buku"""
    )
    status_penerimaan = cur.fetchone()

    # buku dengan stok menipis
    cur.execute(
        """SELECT * FROM buku 
           WHERE stok_minimum > 0 AND stok <= stok_minimum 
           ORDER BY stok ASC LIMIT 10"""
    )
    stok_menipis = cur.fetchall()

    # transaksi hari ini (total eksemplar masuk & keluar)
    cur.execute(
        """SELECT tipe, COUNT(*) as jumlah_transaksi, COALESCE(SUM(jumlah), 0) as total_item
           FROM transaksi
           WHERE tanggal = CURRENT_DATE
           GROUP BY tipe"""
    )
    transaksi_hari_ini_raw = cur.fetchall()
    transaksi_hari_ini = {'masuk': {'jumlah_transaksi': 0, 'total_item': 0},
                           'keluar': {'jumlah_transaksi': 0, 'total_item': 0}}
    for row in transaksi_hari_ini_raw:
        transaksi_hari_ini[row['tipe']] = row

    # jumlah judul buku unik yang masuk hari ini
    cur.execute(
        """SELECT COUNT(DISTINCT buku_id) as total_judul_masuk
           FROM transaksi
           WHERE tipe = 'masuk' AND tanggal = CURRENT_DATE"""
    )
    judul_masuk_hari_ini = cur.fetchone()['total_judul_masuk']

    # 5 transaksi terakhir
    cur.execute(
        """SELECT t.*, b.judul
           FROM transaksi t
           JOIN buku b ON t.buku_id = b.id
           ORDER BY t.created_at DESC
           LIMIT 5"""
    )
    aktivitas_terbaru = cur.fetchall()
    # ringkasan progress distribusi ke tujuan
    cur.execute(
        """SELECT 
             COUNT(DISTINCT t.id) as total_tujuan,
             COALESCE(SUM(dr.jumlah_rencana), 0) as total_rencana_distribusi
           FROM tujuan t
           LEFT JOIN distribusi_rencana dr ON dr.tujuan_id = t.id"""
    )
    ringkasan_distribusi = cur.fetchone()

    cur.execute(
        "SELECT COALESCE(SUM(jumlah), 0) as total FROM transaksi WHERE tipe = 'keluar' AND tujuan_id IS NOT NULL"
    )
    total_terkirim_distribusi = cur.fetchone()['total']

    cur.execute(
        """SELECT 
             t.id,
             COALESCE(SUM(dr.jumlah_rencana), 0) as rencana,
             COALESCE((
                 SELECT SUM(tr.jumlah) FROM transaksi tr 
                 WHERE tr.tujuan_id = t.id AND tr.tipe = 'keluar'
             ), 0) as terkirim
           FROM tujuan t
           LEFT JOIN distribusi_rencana dr ON dr.tujuan_id = t.id
           GROUP BY t.id"""
    )
    semua_tujuan_ringkas = cur.fetchall()

    tujuan_belum_ada_rencana = 0
    tujuan_belum_dikirim = 0
    tujuan_sebagian = 0
    tujuan_lengkap = 0
    for t in semua_tujuan_ringkas:
        if t['rencana'] == 0:
            tujuan_belum_ada_rencana += 1
        elif t['terkirim'] == 0:
            tujuan_belum_dikirim += 1
        elif t['terkirim'] < t['rencana']:
            tujuan_sebagian += 1
        else:
            tujuan_lengkap += 1
    cur.close()
    conn.close()

    return render_template(
        'dashboard.html',
        ringkasan=ringkasan,
        stok_menipis=stok_menipis,
        transaksi_hari_ini=transaksi_hari_ini,
        aktivitas_terbaru=aktivitas_terbaru,
        status_penerimaan=status_penerimaan,
        judul_masuk_hari_ini=judul_masuk_hari_ini,
        ringkasan_distribusi=ringkasan_distribusi,
        total_terkirim_distribusi=total_terkirim_distribusi,
        tujuan_belum_ada_rencana=tujuan_belum_ada_rencana,
        tujuan_belum_dikirim=tujuan_belum_dikirim,
        tujuan_sebagian=tujuan_sebagian,
        tujuan_lengkap=tujuan_lengkap
    )

# ------------------ LIST BUKU ------------------
@app.route('/buku')
@login_required
def buku_list():
    search = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '').strip()
    penerbit_filter = request.args.get('penerbit', '').strip()
    tgl_dari = request.args.get('tgl_dari', '').strip()
    tgl_sampai = request.args.get('tgl_sampai', '').strip()
    catatan_filter = request.args.get('catatan', '').strip()
    catatan_filter = request.args.get('catatan', '').strip()

    conn = get_db_connection()
    cur = conn.cursor()

    query = "SELECT * FROM buku WHERE 1=1"
    params = []

    if search:
        query += " AND (judul ILIKE %s OR isbn ILIKE %s OR penulis ILIKE %s)"
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])

    if status_filter == 'belum':
        query += " AND jumlah_rencana > 0 AND stok = 0"
    elif status_filter == 'sebagian':
        query += " AND jumlah_rencana > 0 AND stok > 0 AND stok < jumlah_rencana"
    elif status_filter == 'lengkap':
        query += " AND jumlah_rencana > 0 AND stok = jumlah_rencana"
    elif status_filter == 'lebih':
        query += " AND jumlah_rencana > 0 AND stok > jumlah_rencana"

    if penerbit_filter:
        query += " AND penerbit = %s"
        params.append(penerbit_filter)

    if tgl_dari:
        query += " AND tanggal_masuk >= %s"
        params.append(tgl_dari)

    if tgl_sampai:
        query += " AND tanggal_masuk <= %s"
        params.append(tgl_sampai)

    if catatan_filter == 'kosong':
        query += " AND (catatan IS NULL OR catatan = '')"
    elif catatan_filter == 'isi':
        query += " AND catatan IS NOT NULL AND catatan != ''"

    query += " ORDER BY judul ASC"

    cur.execute(query, tuple(params))
    daftar_buku = cur.fetchall()

    # daftar penerbit unik untuk dropdown filter
    cur.execute("SELECT DISTINCT penerbit FROM buku WHERE penerbit IS NOT NULL AND penerbit != '' ORDER BY penerbit ASC")
    daftar_penerbit = [row['penerbit'] for row in cur.fetchall()]

    cur.close()
    conn.close()

    return render_template(
        'buku/list.html',
        daftar_buku=daftar_buku,
        search=search,
        status_filter=status_filter,
        penerbit_filter=penerbit_filter,
        daftar_penerbit=daftar_penerbit,
        tgl_dari=tgl_dari,
        tgl_sampai=tgl_sampai,
        catatan_filter=catatan_filter
    )

# ------------------ EXPORT BUKU - EXCEL ------------------
@app.route('/buku/export/excel')
@login_required
def buku_export_excel():
    search = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '').strip()
    penerbit_filter = request.args.get('penerbit', '').strip()
    tgl_dari = request.args.get('tgl_dari', '').strip()
    tgl_sampai = request.args.get('tgl_sampai', '').strip()
    catatan_filter = request.args.get('catatan', '').strip()

    conn = get_db_connection()
    cur = conn.cursor()

    query = "SELECT * FROM buku WHERE 1=1"
    params = []

    if search:
        query += " AND (judul ILIKE %s OR isbn ILIKE %s OR penulis ILIKE %s)"
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])

    if status_filter == 'belum':
        query += " AND jumlah_rencana > 0 AND stok = 0"
    elif status_filter == 'sebagian':
        query += " AND jumlah_rencana > 0 AND stok > 0 AND stok < jumlah_rencana"
    elif status_filter == 'lengkap':
        query += " AND jumlah_rencana > 0 AND stok = jumlah_rencana"
    elif status_filter == 'lebih':
        query += " AND jumlah_rencana > 0 AND stok > jumlah_rencana"

    if penerbit_filter:
        query += " AND penerbit = %s"
        params.append(penerbit_filter)

    if tgl_dari:
        query += " AND tanggal_masuk >= %s"
        params.append(tgl_dari)

    if tgl_sampai:
        query += " AND tanggal_masuk <= %s"
        params.append(tgl_sampai)

    if catatan_filter == 'kosong':
        query += " AND (catatan IS NULL OR catatan = '')"
    elif catatan_filter == 'isi':
        query += " AND catatan IS NOT NULL AND catatan != ''"

    query += " ORDER BY judul ASC"

    cur.execute(query, tuple(params))
    daftar_buku = cur.fetchall()
    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Buku"

    headers = ['No', 'ISBN', 'Judul', 'Penulis', 'Penerbit', 'Diterima', 'Rencana', 'Stok Minimum', 'Tgl Masuk']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')

    from openpyxl.styles import Alignment
    wrap_style = Alignment(wrap_text=True, vertical='top')

    for i, buku in enumerate(daftar_buku, start=1):
        ws.append([
            i, buku['isbn'], buku['judul'], buku['penulis'] or '-',
            buku['penerbit'] or '-',
            buku['stok'], buku['jumlah_rencana'], buku['stok_minimum'],
            str(buku['tanggal_masuk']) if buku['tanggal_masuk'] else '-'
        ])
        row_num = ws.max_row
        ws.cell(row=row_num, column=3).alignment = wrap_style  # Judul
        ws.cell(row=row_num, column=4).alignment = wrap_style  # Penulis
        ws.cell(row=row_num, column=5).alignment = wrap_style  # Penerbit

    lebar_kolom = {'A': 6, 'B': 16, 'C': 45, 'D': 30, 'E': 25, 'F': 10, 'G': 10, 'H': 13, 'I': 13}
    for kolom, lebar in lebar_kolom.items():
        ws.column_dimensions[kolom].width = lebar

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    label_status = {'belum': 'belum-diterima', 'sebagian': 'sebagian-diterima', 'lengkap': 'lengkap-diterima', 'lebih': 'lebih-diterima'}.get(status_filter, 'semua')
    filename = f"data-buku-{label_status}-{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
# ------------------ EXPORT BUKU - PDF ------------------
@app.route('/buku/export/pdf')
@login_required
def buku_export_pdf():
    search = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '').strip()
    penerbit_filter = request.args.get('penerbit', '').strip()
    tgl_dari = request.args.get('tgl_dari', '').strip()
    tgl_sampai = request.args.get('tgl_sampai', '').strip()

    conn = get_db_connection()
    cur = conn.cursor()

    query = "SELECT * FROM buku WHERE 1=1"
    params = []

    if search:
        query += " AND (judul ILIKE %s OR isbn ILIKE %s OR penulis ILIKE %s)"
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])

    if status_filter == 'belum':
        query += " AND jumlah_rencana > 0 AND stok = 0"
    elif status_filter == 'sebagian':
        query += " AND jumlah_rencana > 0 AND stok > 0 AND stok < jumlah_rencana"
    elif status_filter == 'lengkap':
        query += " AND jumlah_rencana > 0 AND stok >= jumlah_rencana"

    if penerbit_filter:
        query += " AND penerbit = %s"
        params.append(penerbit_filter)

    if tgl_dari:
        query += " AND tanggal_masuk >= %s"
        params.append(tgl_dari)

    if tgl_sampai:
        query += " AND tanggal_masuk <= %s"
        params.append(tgl_sampai)

    if catatan_filter == 'kosong':
        query += " AND (catatan IS NULL OR catatan = '')"
    elif catatan_filter == 'isi':
        query += " AND catatan IS NOT NULL AND catatan != ''"

    query += " ORDER BY judul ASC"

    cur.execute(query, tuple(params))
    daftar_buku = cur.fetchall()
    cur.close()
    conn.close()

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=landscape(A4),
        topMargin=1*cm, bottomMargin=1*cm, leftMargin=1*cm, rightMargin=1*cm
    )
    styles = getSampleStyleSheet()
    elements = []

    judul_map = {'belum': ' - Belum Diterima', 'sebagian': ' - Diterima Sebagian', 'lengkap': ' - Sudah Lengkap'}
    judul_laporan = "Laporan Data Buku" + judul_map.get(status_filter, '')
    elements.append(Paragraph(judul_laporan, styles['Title']))
    elements.append(Paragraph(f"Dicetak: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    style_sel = ParagraphStyle('sel', fontSize=7.5, leading=9, fontName='Helvetica')
    style_header = ParagraphStyle('header', fontSize=8, leading=10, fontName='Helvetica-Bold', textColor=colors.white)

    data = [[
        Paragraph('No', style_header), Paragraph('ISBN', style_header), Paragraph('Judul', style_header),
        Paragraph('Penulis', style_header), Paragraph('Penerbit', style_header),
        Paragraph('Diterima', style_header), Paragraph('Rencana', style_header),
        Paragraph('Min', style_header), Paragraph('Tgl Masuk', style_header)
    ]]
    for i, buku in enumerate(daftar_buku, start=1):
        data.append([
            Paragraph(str(i), style_sel),
            Paragraph(buku['isbn'], style_sel),
            Paragraph(buku['judul'], style_sel),
            Paragraph(buku['penulis'] or '-', style_sel),
            Paragraph(buku['penerbit'] or '-', style_sel),
            Paragraph(str(buku['stok']), style_sel),
            Paragraph(str(buku['jumlah_rencana']), style_sel),
            Paragraph(str(buku['stok_minimum']), style_sel),
            Paragraph(str(buku['tanggal_masuk']) if buku['tanggal_masuk'] else '-', style_sel),
        ])

    col_widths = [1.0*cm, 2.4*cm, 6.5*cm, 5.0*cm, 4.0*cm, 1.8*cm, 1.8*cm, 1.3*cm, 2.2*cm]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D6EFD')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)

    doc.build(elements)
    output.seek(0)

    label_status = {'belum': 'belum-diterima', 'sebagian': 'sebagian-diterima', 'lengkap': 'lengkap-diterima'}.get(status_filter, 'semua')
    filename = f"data-buku-{label_status}-{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(output, mimetype='application/pdf', as_attachment=True, download_name=filename)
# ------------------ EXPORT TRANSAKSI - EXCEL ------------------
@app.route('/transaksi/export/excel')
@login_required
def transaksi_export_excel():
    tipe_filter = request.args.get('tipe', '').strip()

    conn = get_db_connection()
    cur = conn.cursor()
    query = """
        SELECT t.*, b.judul, b.isbn, b.penerbit, u.nama_lengkap, u.username
        FROM transaksi t
        JOIN buku b ON t.buku_id = b.id
        LEFT JOIN users u ON t.user_id = u.id
    """
    params = []
    if tipe_filter in ('masuk', 'keluar'):
        query += " WHERE t.tipe = %s"
        params.append(tipe_filter)
    query += " ORDER BY t.tanggal DESC, t.created_at DESC"
    cur.execute(query, tuple(params))
    daftar_transaksi = cur.fetchall()
    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Riwayat Transaksi"

    headers = ['Tanggal', 'Tipe', 'ISBN', 'Judul', 'Penerbit', 'Jumlah', 'Pihak Terkait', 'Keterangan', 'Dicatat Oleh']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')

    for t in daftar_transaksi:
        ws.append([
            str(t['tanggal']), t['tipe'].capitalize(), t['isbn'], t['judul'], t['penerbit'] or '-',
            t['jumlah'], t['pihak_terkait'] or '-', t['keterangan'] or '-',
            t['nama_lengkap'] or t['username'] or '-'
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_length + 3

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"riwayat-transaksi-{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ------------------ EXPORT TRANSAKSI - PDF ------------------
@app.route('/transaksi/export/pdf')
@login_required
def transaksi_export_pdf():
    tipe_filter = request.args.get('tipe', '').strip()

    conn = get_db_connection()
    cur = conn.cursor()
    query = """
        SELECT t.*, b.judul, b.isbn, b.penerbit, u.nama_lengkap, u.username
        FROM transaksi t
        JOIN buku b ON t.buku_id = b.id
        LEFT JOIN users u ON t.user_id = u.id
    """
    params = []
    if tipe_filter in ('masuk', 'keluar'):
        query += " WHERE t.tipe = %s"
        params.append(tipe_filter)
    query += " ORDER BY t.tanggal DESC, t.created_at DESC"
    cur.execute(query, tuple(params))
    daftar_transaksi = cur.fetchall()
    cur.close()
    conn.close()

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    judul_laporan = "Laporan Riwayat Transaksi"
    if tipe_filter == 'masuk':
        judul_laporan += " - Barang Masuk"
    elif tipe_filter == 'keluar':
        judul_laporan += " - Barang Keluar"

    elements.append(Paragraph(judul_laporan, styles['Title']))
    elements.append(Paragraph(f"Dicetak: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    data = [['Tanggal', 'Tipe', 'ISBN', 'Judul', 'Penerbit', 'Jml', 'Pihak Terkait', 'Keterangan', 'Oleh']]
    for t in daftar_transaksi:
        data.append([
            str(t['tanggal']), t['tipe'].capitalize(), t['isbn'], t['judul'], t['penerbit'] or '-',
            str(t['jumlah']), t['pihak_terkait'] or '-', t['keterangan'] or '-',
            t['nama_lengkap'] or t['username'] or '-'
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D6EFD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)

    doc.build(elements)
    output.seek(0)

    filename = f"riwayat-transaksi-{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(output, mimetype='application/pdf', as_attachment=True, download_name=filename)
# ------------------ TAMBAH BUKU ------------------
@app.route('/buku/tambah', methods=['GET', 'POST'])
@login_required
@admin_required 
def buku_tambah():
    if request.method == 'POST':
        isbn = request.form.get('isbn', '').strip()
        judul = request.form.get('judul', '').strip()
        penulis = request.form.get('penulis', '').strip()
        penerbit = request.form.get('penerbit', '').strip()
        
        stok = ambil_int(request.form, 'stok', 0)
        stok_minimum = ambil_int(request.form, 'stok_minimum', 0)
        jumlah_rencana = ambil_int(request.form, 'jumlah_rencana', 0)
        tanggal_masuk = request.form.get('tanggal_masuk', '').strip() or None
        catatan = request.form.get('catatan', '').strip()

        if not isbn or not judul:
            flash('ISBN dan Judul wajib diisi.', 'danger')
            return render_template('buku/form.html', buku=request.form)

        conn = get_db_connection()
        cur = conn.cursor()

        # cek ISBN sudah ada atau belum
        cur.execute("SELECT id FROM buku WHERE isbn = %s", (isbn,))
        existing = cur.fetchone()
        if existing:
            flash(f'ISBN {isbn} sudah terdaftar di database.', 'danger')
            cur.close()
            conn.close()
            return render_template('buku/form.html', buku=request.form)

        cur.execute(
            """INSERT INTO buku (isbn, judul, penulis, penerbit, stok, stok_minimum, jumlah_rencana, tanggal_masuk, catatan)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (isbn, judul, penulis, penerbit, stok, stok_minimum, jumlah_rencana, tanggal_masuk, catatan)
        )
        conn.commit()
        cur.close()
        conn.close()
        catat_aktivitas('Menambah Buku', f'Buku "{judul}" (ISBN {isbn}) ditambahkan')
        flash(f'Buku "{judul}" berhasil ditambahkan.', 'success')
        return redirect(url_for('buku_list'))

    return render_template('buku/form.html', buku=None)


# ------------------ EDIT BUKU ------------------
@app.route('/buku/edit/<int:buku_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def buku_edit(buku_id):
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == 'POST':
        isbn = request.form.get('isbn', '').strip()
        judul = request.form.get('judul', '').strip()
        penulis = request.form.get('penulis', '').strip()
        penerbit = request.form.get('penerbit', '').strip()
        stok = ambil_int(request.form, 'stok', 0)
        stok_minimum = ambil_int(request.form, 'stok_minimum', 0)
        jumlah_rencana = ambil_int(request.form, 'jumlah_rencana', 0)
        tanggal_masuk = request.form.get('tanggal_masuk', '').strip() or None
        catatan = request.form.get('catatan', '').strip()

        if not isbn or not judul:
            flash('ISBN dan Judul wajib diisi.', 'danger')
            cur.close()
            conn.close()
            return render_template('buku/form.html', buku=request.form, buku_id=buku_id)

        cur.execute("SELECT * FROM buku WHERE id = %s", (buku_id,))
        buku_lama = cur.fetchone()

        perubahan = []
        if buku_lama:
            if buku_lama['judul'] != judul:
                perubahan.append(f"judul: \"{buku_lama['judul']}\" → \"{judul}\"")
            if buku_lama['penerbit'] != penerbit:
                perubahan.append(f"penerbit: \"{buku_lama['penerbit'] or '-'}\" → \"{penerbit or '-'}\"")
            if str(buku_lama['stok']) != str(stok):
                perubahan.append(f"stok: {buku_lama['stok']} → {stok}")
            if str(buku_lama['jumlah_rencana']) != str(jumlah_rencana):
                perubahan.append(f"rencana: {buku_lama['jumlah_rencana']} → {jumlah_rencana}")

        cur.execute(
            """UPDATE buku 
               SET isbn=%s, judul=%s, penulis=%s, penerbit=%s, 
                   stok=%s, stok_minimum=%s, jumlah_rencana=%s, tanggal_masuk=%s, catatan=%s, updated_at=NOW()
               WHERE id=%s""",
            (isbn, judul, penulis, penerbit, stok, stok_minimum, jumlah_rencana, tanggal_masuk, catatan, buku_id)
        )
        conn.commit()
        cur.close()
        conn.close()

        detail_perubahan = '; '.join(perubahan) if perubahan else 'tidak ada perubahan data'
        catat_aktivitas('Mengedit Buku', f'Buku "{judul}": {detail_perubahan}', buku_id=buku_id)

        flash(f'Buku "{judul}" berhasil diupdate.', 'success')
        return redirect(url_for('buku_list'))

    cur.execute("SELECT * FROM buku WHERE id = %s", (buku_id,))
    buku = cur.fetchone()
    cur.close()
    conn.close()

    if not buku:
        flash('Buku tidak ditemukan.', 'danger')
        return redirect(url_for('buku_list'))

    return render_template('buku/form.html', buku=buku, buku_id=buku_id)

# ------------------ DETAIL BUKU + RIWAYAT MUTASI ------------------
@app.route('/buku/<int:buku_id>/detail')
@login_required
def buku_detail(buku_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM buku WHERE id = %s", (buku_id,))
    buku = cur.fetchone()

    if not buku:
        cur.close()
        conn.close()
        flash('Buku tidak ditemukan.', 'danger')
        return redirect(url_for('buku_list'))

        cur.execute(
        """SELECT t.*, u.nama_lengkap, u.username
           FROM transaksi t
           LEFT JOIN users u ON t.user_id = u.id
           WHERE t.buku_id = %s
           ORDER BY t.tanggal DESC, t.created_at DESC""",
        (buku_id,)
    )
    riwayat = cur.fetchall()

    cur.execute(
        "SELECT * FROM activity_log WHERE buku_id = %s ORDER BY created_at DESC LIMIT 20",
        (buku_id,)
    )
    riwayat_edit = cur.fetchall()

    # ringkasan total masuk & keluar sepanjang waktu untuk buku ini
    cur.execute(
        """SELECT tipe, COALESCE(SUM(jumlah), 0) as total
           FROM transaksi WHERE buku_id = %s GROUP BY tipe""",
        (buku_id,)
    )
    ringkasan_raw = cur.fetchall()
    ringkasan = {'masuk': 0, 'keluar': 0}
    for row in ringkasan_raw:
        ringkasan[row['tipe']] = row['total']

    cur.close()
    conn.close()

    return render_template('buku/detail.html', buku=buku, riwayat=riwayat, ringkasan=ringkasan, riwayat_edit=riwayat_edit)

# ------------------ HAPUS BUKU ------------------
@app.route('/buku/hapus/<int:buku_id>', methods=['POST'])
@login_required
@admin_required
def buku_hapus(buku_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT judul FROM buku WHERE id = %s", (buku_id,))
    buku = cur.fetchone()

    if not buku:
        flash('Buku tidak ditemukan.', 'danger')
    else:
        try:
            cur.execute("DELETE FROM buku WHERE id = %s", (buku_id,))
            conn.commit()
            catat_aktivitas('Menghapus Buku', f'Buku "{buku["judul"]}" dihapus')
            flash(f'Buku "{buku["judul"]}" berhasil dihapus.', 'success')
        except Exception as e:
            conn.rollback()
            flash('Gagal menghapus buku — mungkin masih ada transaksi terkait.', 'danger')

    cur.close()
    conn.close()
    return redirect(url_for('buku_list'))

# ------------------ BARANG MASUK ------------------
@app.route('/transaksi/masuk', methods=['GET', 'POST'])
@login_required
@viewer_blocked
def transaksi_masuk():
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == 'POST':
        buku_id = request.form.get('buku_id', '').strip()
        jumlah = ambil_int(request.form, 'jumlah', 0)
        catatan_buku = request.form.get('catatan_buku', '').strip()
        pihak_terkait = request.form.get('pihak_terkait', '').strip()
        tanggal = request.form.get('tanggal', '').strip()

        if not buku_id or jumlah <= 0:
            flash('Buku dan jumlah (harus lebih dari 0) wajib diisi.', 'danger')
            cur.execute("SELECT * FROM buku ORDER BY judul ASC")
            daftar_buku = cur.fetchall()
            cur.close()
            conn.close()
            return render_template('transaksi/masuk.html', daftar_buku=daftar_buku, today=datetime.now().date().isoformat())

        try:
            # cek buku ada
            cur.execute("SELECT * FROM buku WHERE id = %s", (buku_id,))
            buku = cur.fetchone()
            if not buku:
                flash('Buku tidak ditemukan.', 'danger')
                cur.close()
                conn.close()
                return redirect(url_for('transaksi_masuk'))

            # insert transaksi
            cur.execute(
                """INSERT INTO transaksi (buku_id, tipe, jumlah, pihak_terkait, user_id, tanggal)
                   VALUES (%s, 'masuk', %s, %s, %s, %s)""",
                (buku_id, jumlah, pihak_terkait, session['user_id'], tanggal or None)
            )

            # update stok buku + tanggal masuk terakhir + catatan
            tgl_transaksi = tanggal or datetime.now().date()
            cur.execute(
                """UPDATE buku 
                   SET stok = stok + %s, updated_at = NOW(),
                       tanggal_masuk = GREATEST(COALESCE(tanggal_masuk, %s), %s),
                       catatan = %s
                   WHERE id = %s""",
                (jumlah, tgl_transaksi, tgl_transaksi, catatan_buku, buku_id)
            )

            conn.commit()
            flash(f'Barang masuk: {buku["judul"]} +{jumlah} berhasil dicatat.', 'success')
        except Exception as e:
            conn.rollback()
            flash('Gagal mencatat transaksi. Coba lagi.', 'danger')
        finally:
            cur.close()
            conn.close()

        return redirect(url_for('transaksi_masuk'))

    cur.execute("SELECT * FROM buku ORDER BY judul ASC")
    daftar_buku = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('transaksi/masuk.html', daftar_buku=daftar_buku, today=datetime.now().date().isoformat())


# ------------------ BARANG KELUAR ------------------
@app.route('/transaksi/keluar', methods=['GET', 'POST'])
@login_required
@viewer_blocked
def transaksi_keluar():
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == 'POST':
        buku_id = request.form.get('buku_id', '').strip()
        jumlah = ambil_int(request.form, 'jumlah', 0)
        keterangan = request.form.get('keterangan', '').strip()
        pihak_terkait = request.form.get('pihak_terkait', '').strip()
        tanggal = request.form.get('tanggal', '').strip()
        tujuan_id = request.form.get('tujuan_id', '').strip() or None
        if not buku_id or jumlah <= 0:
            flash('Buku dan jumlah (harus lebih dari 0) wajib diisi.', 'danger')
            cur.execute("SELECT * FROM buku ORDER BY judul ASC")
            daftar_buku = cur.fetchall()
            cur.close()
            conn.close()
            return render_template('transaksi/keluar.html', daftar_buku=daftar_buku)

        try:
            cur.execute("SELECT * FROM buku WHERE id = %s", (buku_id,))
            buku = cur.fetchone()

            if not buku:
                flash('Buku tidak ditemukan.', 'danger')
                cur.close()
                conn.close()
                return redirect(url_for('transaksi_keluar'))

            if buku['stok'] < jumlah:
                flash(f'Stok tidak cukup. Stok tersedia: {buku["stok"]}, diminta: {jumlah}.', 'danger')
                cur.close()
                conn.close()
                return redirect(url_for('transaksi_keluar'))

            cur.execute(
                """INSERT INTO transaksi (buku_id, tipe, jumlah, keterangan, pihak_terkait, user_id, tanggal, tujuan_id)
                   VALUES (%s, 'keluar', %s, %s, %s, %s, %s, %s)""",
                (buku_id, jumlah, keterangan, pihak_terkait, session['user_id'], tanggal or None, tujuan_id)
            )

            cur.execute(
                "UPDATE buku SET stok = stok - %s, updated_at = NOW() WHERE id = %s",
                (jumlah, buku_id)
            )

            conn.commit()
            flash(f'Barang keluar: {buku["judul"]} -{jumlah} berhasil dicatat.', 'success')
        except Exception as e:
            conn.rollback()
            flash('Gagal mencatat transaksi. Coba lagi.', 'danger')
        finally:
            cur.close()
            conn.close()

        return redirect(url_for('transaksi_keluar'))

    cur.execute("SELECT * FROM buku ORDER BY judul ASC")
    daftar_buku = cur.fetchall()
    cur.execute("SELECT * FROM tujuan ORDER BY nama ASC")
    daftar_tujuan = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('transaksi/keluar.html', daftar_buku=daftar_buku, daftar_tujuan=daftar_tujuan)


# ------------------ RIWAYAT TRANSAKSI ------------------
@app.route('/transaksi/riwayat')
@login_required
@viewer_blocked
def transaksi_riwayat():
    tipe_filter = request.args.get('tipe', '').strip()

    conn = get_db_connection()
    cur = conn.cursor()

    query = """
        SELECT t.*, b.judul, b.isbn, b.penerbit, u.nama_lengkap, u.username
        FROM transaksi t
        JOIN buku b ON t.buku_id = b.id
        LEFT JOIN users u ON t.user_id = u.id
    """
    params = []

    if tipe_filter in ('masuk', 'keluar'):
        query += " WHERE t.tipe = %s"
        params.append(tipe_filter)

    query += " ORDER BY t.tanggal DESC, t.created_at DESC"

    cur.execute(query, tuple(params))
    daftar_transaksi = cur.fetchall()
    cur.close()
    conn.close()

    return render_template('transaksi/riwayat.html', daftar_transaksi=daftar_transaksi, tipe_filter=tipe_filter)

# ------------------ API: LOOKUP ISBN (Google Books) ------------------
@app.route('/api/isbn/<isbn>')
@login_required
def api_lookup_isbn(isbn):
    isbn = isbn.strip().replace('-', '').replace(' ', '')

    try:
        resp = requests.get(
            f'https://www.googleapis.com/books/v1/volumes?q=isbn:{isbn}',
            timeout=5
        )
        data = resp.json()

        if data.get('totalItems', 0) > 0:
            info = data['items'][0]['volumeInfo']
            return {
                'found': True,
                'judul': info.get('title', ''),
                'penulis': ', '.join(info.get('authors', [])),
                'penerbit': info.get('publisher', ''),
                'sampul_url': info.get('imageLinks', {}).get('thumbnail', '')
            }
        else:
            return {'found': False}
    except Exception as e:
        return {'found': False, 'error': str(e)}
    # ------------------ API: GENERATE KODE INTERNAL BARU ------------------
@app.route('/api/kode-internal-baru')
@login_required
def api_kode_internal_baru():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT isbn FROM buku WHERE isbn LIKE 'GDG-%' ORDER BY isbn DESC LIMIT 1")
    last = cur.fetchone()
    cur.close()
    conn.close()

    last_num = 0
    if last:
        try:
            last_num = int(last['isbn'].split('-')[1])
        except (IndexError, ValueError):
            last_num = 0

    return {'kode': f"GDG-{last_num + 1:04d}"}


# ------------------ GENERATE GAMBAR BARCODE ------------------
@app.route('/buku/<int:buku_id>/barcode.png')
@login_required
def buku_barcode_image(buku_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT isbn FROM buku WHERE id = %s", (buku_id,))
    buku = cur.fetchone()
    cur.close()
    conn.close()

    if not buku:
        return '', 404

    code128 = barcode.get('code128', buku['isbn'], writer=ImageWriter())
    output = io.BytesIO()
    code128.write(output, options={'write_text': False, 'module_height': 8.0, 'quiet_zone': 2})
    output.seek(0)
    return send_file(output, mimetype='image/png')


# ------------------ HALAMAN PILIH BUKU UNTUK CETAK LABEL ------------------
@app.route('/buku/cetak-label', methods=['GET', 'POST'])
@login_required
def buku_cetak_label():
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == 'POST':
        ids = request.form.getlist('buku_ids')
        ids = [int(i) for i in ids if i.isdigit()]

        if not ids:
            flash('Pilih minimal satu buku untuk dicetak labelnya.', 'danger')
            cur.execute("SELECT * FROM buku ORDER BY judul ASC")
            daftar_buku = cur.fetchall()
            cur.close()
            conn.close()
            return render_template('buku/pilih_label.html', daftar_buku=daftar_buku)

        cur.execute("SELECT * FROM buku WHERE id = ANY(%s) ORDER BY judul ASC", (ids,))
        terpilih = cur.fetchall()
        cur.close()
        conn.close()
        return render_template('buku/label.html', daftar_buku=terpilih)

    cur.execute("SELECT * FROM buku ORDER BY judul ASC")
    daftar_buku = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('buku/pilih_label.html', daftar_buku=daftar_buku)


# ------------------ CETAK LABEL UNTUK 1 BUKU (dari halaman edit) ------------------
@app.route('/buku/<int:buku_id>/label')
@login_required
def buku_label(buku_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM buku WHERE id = %s", (buku_id,))
    buku = cur.fetchone()
    cur.close()
    conn.close()

    if not buku:
        flash('Buku tidak ditemukan.', 'danger')
        return redirect(url_for('buku_list'))

    return render_template('buku/label.html', daftar_buku=[buku])

# ------------------ DAFTAR PEMINJAMAN AKTIF ------------------
@app.route('/peminjaman')
@login_required
@viewer_blocked
def peminjaman_list():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """SELECT t.*, b.judul, b.isbn
           FROM transaksi t
           JOIN buku b ON t.buku_id = b.id
           WHERE t.tipe = 'keluar' AND t.jenis_keluar = 'pinjam' 
                 AND t.tanggal_kembali_aktual IS NULL
           ORDER BY t.tanggal_kembali_rencana ASC"""
    )
    daftar_pinjam = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('peminjaman.html', daftar_pinjam=daftar_pinjam, today=datetime.now().date())


# ------------------ TANDAI SUDAH KEMBALI ------------------
@app.route('/peminjaman/<int:transaksi_id>/kembali', methods=['POST'])
@login_required
@viewer_blocked
def peminjaman_kembali(transaksi_id):
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute(
            """SELECT * FROM transaksi 
               WHERE id = %s AND tipe = 'keluar' AND jenis_keluar = 'pinjam' 
                     AND tanggal_kembali_aktual IS NULL""",
            (transaksi_id,)
        )
        transaksi = cur.fetchone()

        if not transaksi:
            flash('Data peminjaman tidak ditemukan atau sudah ditandai kembali.', 'danger')
        else:
            cur.execute(
                "UPDATE transaksi SET tanggal_kembali_aktual = CURRENT_DATE WHERE id = %s",
                (transaksi_id,)
            )
            cur.execute(
                "UPDATE buku SET stok = stok + %s, updated_at = NOW() WHERE id = %s",
                (transaksi['jumlah'], transaksi['buku_id'])
            )
            conn.commit()
            flash('Berhasil ditandai sudah kembali, stok bertambah kembali.', 'success')
    except Exception as e:
        conn.rollback()
        flash('Gagal memproses. Coba lagi.', 'danger')
    finally:
        cur.close()
        conn.close()

    return redirect(url_for('peminjaman_list'))
# ------------------ ADMIN: DAFTAR USER ------------------
@app.route('/admin/users')
@login_required
@admin_required
def user_list():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM users ORDER BY username ASC")
    daftar_user = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('admin/users.html', daftar_user=daftar_user)


# ------------------ ADMIN: TAMBAH USER ------------------
@app.route('/admin/users/tambah', methods=['GET', 'POST'])
@login_required
@admin_required
def user_tambah():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        nama_lengkap = request.form.get('nama_lengkap', '').strip()
        role = request.form.get('role', 'staff').strip()

        if not username or not password:
            flash('Username dan password wajib diisi.', 'danger')
            return render_template('admin/user_form.html')

        if len(password) < 8:
            flash('Password minimal 8 karakter.', 'danger')
            return render_template('admin/user_form.html')

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("SELECT id FROM users WHERE username = %s", (username,))
        if cur.fetchone():
            flash('Username sudah dipakai.', 'danger')
            cur.close()
            conn.close()
            return render_template('admin/user_form.html')

        password_hash = generate_password_hash(password)
        cur.execute(
            """INSERT INTO users (username, password_hash, nama_lengkap, role)
               VALUES (%s, %s, %s, %s)""",
            (username, password_hash, nama_lengkap, role)
        )
        conn.commit()
        cur.close()
        conn.close()
        catat_aktivitas('Menambah User', f'User baru "{username}" (role: {role}) ditambahkan')
        flash(f'User "{username}" berhasil ditambahkan.', 'success')
        return redirect(url_for('user_list'))

    return render_template('admin/user_form.html')


# ------------------ ADMIN: AKTIFKAN/NONAKTIFKAN USER ------------------
@app.route('/admin/users/<int:user_id>/toggle', methods=['POST'])
@login_required
@admin_required
def user_toggle(user_id):
    if user_id == session['user_id']:
        flash('Tidak bisa menonaktifkan akun sendiri.', 'danger')
        return redirect(url_for('user_list'))

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT username, is_active FROM users WHERE id = %s", (user_id,))
    user = cur.fetchone()

    if user:
        status_baru = not user['is_active']
        cur.execute("UPDATE users SET is_active = %s WHERE id = %s", (status_baru, user_id))
        conn.commit()
        flash('Status user berhasil diubah.', 'success')
        aksi_label = 'Mengaktifkan' if status_baru else 'Menonaktifkan'
        catat_aktivitas(aksi_label, f'User "{user["username"]}" di{"aktifkan" if status_baru else "nonaktifkan"}')

    cur.close()
    conn.close()
    return redirect(url_for('user_list'))


# ------------------ ADMIN: RESET PASSWORD USER ------------------
@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
@admin_required
def user_reset_password(user_id):
    password_baru = request.form.get('password_baru', '').strip()

    if len(password_baru) < 8:
        flash('Password minimal 8 karakter.', 'danger')
        return redirect(url_for('user_list'))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT username FROM users WHERE id = %s", (user_id,))
    target_user = cur.fetchone()

    cur.execute(
        "UPDATE users SET password_hash = %s, failed_attempts = 0, locked_until = NULL WHERE id = %s",
        (generate_password_hash(password_baru), user_id)
    )
    conn.commit()
    cur.close()
    conn.close()

    flash('Password berhasil direset.', 'success')
    catat_aktivitas('Reset Password', f'Password user "{target_user["username"] if target_user else user_id}" direset')
    return redirect(url_for('user_list'))
# ------------------ ADMIN: KIRIM NOTIFIKASI MANUAL ------------------
@app.route('/admin/kirim-notifikasi-stok', methods=['POST'])
@login_required
@admin_required
def kirim_notifikasi_stok_manual():
    sukses, pesan = kirim_email_stok_kritis()
    flash(pesan, 'success' if sukses else 'danger')
    return redirect(url_for('dashboard'))


# ------------------ CRON: DIPANGGIL SCHEDULER EKSTERNAL ------------------
@app.route('/cron/cek-stok-kritis')
def cron_cek_stok_kritis():
    secret = request.args.get('secret', '')
    if secret != os.environ.get('CRON_SECRET'):
        return {'error': 'unauthorized'}, 401

    sukses, pesan = kirim_email_stok_kritis()
    return {'success': sukses, 'message': pesan}

# ------------------ DOWNLOAD TEMPLATE IMPORT ------------------
@app.route('/buku/import/template')
@login_required
@admin_required
def buku_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Import"

    headers = ['isbn', 'judul', 'penulis', 'penerbit', 'jumlah_rencana', 'stok', 'stok_minimum']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')

    ws.append(['9786020633178', 'Contoh Judul Buku', 'Nama Penulis', 'Nama Penerbit', 40, 0, 3])

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_length + 3

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='template-import-buku.xlsx'
    )

# ------------------ IMPORT BUKU DARI EXCEL ------------------
@app.route('/buku/import', methods=['GET', 'POST'])
@login_required
@admin_required
def buku_import():
    if request.method == 'POST':
        file = request.files.get('file_excel')

        if not file or file.filename == '':
            flash('Pilih file Excel dulu.', 'danger')
            return render_template('buku/import.html')

        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('File harus berformat .xlsx atau .xls', 'danger')
            return render_template('buku/import.html')

        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
        except Exception as e:
            flash('Gagal membaca file Excel. Pastikan formatnya benar.', 'danger')
            return render_template('buku/import.html')

        # alias nama kolom yang dikenali (huruf kecil, sudah di-strip)
        alias_kolom = {
            'isbn': ['isbn'],
            'judul': ['judul', 'judul buku'],
            'penulis': ['penulis'],
            'penerbit': ['penerbit'],
            'jumlah_rencana': ['eksemplar', 'jumlah rencana', 'jumlah_rencana', 'rencana'],
            'stok': ['stok', 'stok awal'],
            'stok_minimum': ['stok_minimum', 'stok minimum'],
        }

        # cari baris header: scan 15 baris pertama, cari baris yang punya sel "isbn"
        header_row_num = None
        kolom_index = {}

        for row_num in range(1, min(16, ws.max_row + 1)):
            row_values = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[row_num]]
            if 'isbn' in row_values:
                header_row_num = row_num
                for field, kemungkinan_nama in alias_kolom.items():
                    for nama in kemungkinan_nama:
                        if nama in row_values:
                            kolom_index[field] = row_values.index(nama)
                            break
                break

        if header_row_num is None or 'isbn' not in kolom_index or 'judul' not in kolom_index:
            flash('Kolom ISBN dan Judul tidak ditemukan di file. Pastikan ada baris header dengan kolom "ISBN" dan "Judul"/"Judul Buku".', 'danger')
            return render_template('buku/import.html')

        conn = get_db_connection()
        cur = conn.cursor()

        berhasil = 0
        dilewati = []
        baris_ke = header_row_num

        for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
            baris_ke += 1

            def ambil(nama_field, default=''):
                idx = kolom_index.get(nama_field)
                if idx is None or idx >= len(row) or row[idx] is None:
                    return default
                return row[idx]

            isbn = str(ambil('isbn')).strip()
            judul = str(ambil('judul')).strip()

            if not isbn or not judul or isbn.lower() == 'none':
                continue  # baris kosong, dilewati diam-diam (bukan error)

            cur.execute("SELECT id FROM buku WHERE isbn = %s", (isbn,))
            if cur.fetchone():
                dilewati.append(f"Baris {baris_ke}: ISBN {isbn} sudah ada")
                continue

            try:
                jumlah_rencana = int(float(ambil('jumlah_rencana', 0) or 0))
            except (ValueError, TypeError):
                jumlah_rencana = 0
            try:
                stok = int(float(ambil('stok', 0) or 0))
            except (ValueError, TypeError):
                stok = 0
            try:
                stok_minimum = int(float(ambil('stok_minimum', 0) or 0))
            except (ValueError, TypeError):
                stok_minimum = 0

            cur.execute(
                """INSERT INTO buku (isbn, judul, penulis, penerbit, stok, stok_minimum, jumlah_rencana)
                   VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                (isbn, judul, str(ambil('penulis')), str(ambil('penerbit')),
                 stok, stok_minimum, jumlah_rencana)
            )
            berhasil += 1

        conn.commit()
        cur.close()
        conn.close()

        pesan = f'{berhasil} buku berhasil diimpor.'
        if dilewati:
            pesan += f' {len(dilewati)} baris dilewati (ISBN dobel).'
        flash(pesan, 'success' if berhasil > 0 else 'danger')

        return render_template('buku/import.html', dilewati=dilewati, berhasil=berhasil)

    return render_template('buku/import.html')

# ------------------ DOWNLOAD TEMPLATE IMPORT MASSAL BARANG MASUK ------------------
@app.route('/transaksi/masuk/import/template')
@login_required
@viewer_blocked
def import_masuk_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Barang Masuk"

    headers = ['isbn', 'jumlah_masuk', 'tanggal', 'keterangan']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')

    ws.append(['9786347345264', 40, '2026-08-12', 'Kiriman tahap 1'])

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_length + 3

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='template-barang-masuk-massal.xlsx'
    )
# ------------------ IMPORT MASSAL BARANG MASUK ------------------
@app.route('/transaksi/masuk/import', methods=['GET', 'POST'])
@login_required
@viewer_blocked
@admin_required
def import_masuk_massal():
    if request.method == 'POST':
        file = request.files.get('file_excel')

        if not file or file.filename == '':
            flash('Pilih file Excel dulu.', 'danger')
            return render_template('transaksi/import_masuk.html')

        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('File harus berformat .xlsx atau .xls', 'danger')
            return render_template('transaksi/import_masuk.html')

        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
        except Exception:
            flash('Gagal membaca file Excel. Pastikan formatnya benar.', 'danger')
            return render_template('transaksi/import_masuk.html')

        header_row = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]

        if 'isbn' not in header_row or 'jumlah_masuk' not in header_row:
            flash('Kolom "isbn" dan "jumlah_masuk" wajib ada di baris pertama. Gunakan template yang disediakan.', 'danger')
            return render_template('transaksi/import_masuk.html')

        idx_isbn = header_row.index('isbn')
        idx_jumlah = header_row.index('jumlah_masuk')
        idx_keterangan = header_row.index('keterangan') if 'keterangan' in header_row else None
        idx_tanggal = header_row.index('tanggal') if 'tanggal' in header_row else None

        conn = get_db_connection()
        cur = conn.cursor()

        berhasil = []
        gagal = []
        baris_ke = 1

        for row in ws.iter_rows(min_row=2, values_only=True):
            baris_ke += 1

            isbn = str(row[idx_isbn]).strip() if idx_isbn < len(row) and row[idx_isbn] else ''
            if not isbn:
                continue

            try:
                jumlah = int(float(row[idx_jumlah])) if idx_jumlah < len(row) and row[idx_jumlah] else 0
            except (ValueError, TypeError):
                jumlah = 0

            keterangan = ''
            if idx_keterangan is not None and idx_keterangan < len(row) and row[idx_keterangan]:
                keterangan = str(row[idx_keterangan])

            tgl_masuk = datetime.now().date()
            if idx_tanggal is not None and idx_tanggal < len(row) and row[idx_tanggal]:
                nilai_tanggal = row[idx_tanggal]
                if hasattr(nilai_tanggal, 'date'):
                    tgl_masuk = nilai_tanggal.date()
                elif isinstance(nilai_tanggal, str):
                    try:
                        tgl_masuk = datetime.strptime(nilai_tanggal.strip(), '%Y-%m-%d').date()
                    except ValueError:
                        pass

            if jumlah <= 0:
                gagal.append(f"Baris {baris_ke}: jumlah tidak valid untuk ISBN {isbn}")
                continue

            cur.execute("SELECT * FROM buku WHERE isbn = %s", (isbn,))
            buku = cur.fetchone()

            if not buku:
                gagal.append(f"Baris {baris_ke}: ISBN {isbn} tidak ditemukan di master data")
                continue

            try:
                cur.execute(
                    """INSERT INTO transaksi (buku_id, tipe, jumlah, keterangan, user_id, tanggal)
                       VALUES (%s, 'masuk', %s, %s, %s, %s)""",
                    (buku['id'], jumlah, keterangan or 'Import massal barang masuk', session['user_id'], tgl_masuk)
                )
                cur.execute(
                    """UPDATE buku 
                       SET stok = stok + %s, updated_at = NOW(),
                           tanggal_masuk = GREATEST(COALESCE(tanggal_masuk, %s), %s)
                       WHERE id = %s""",
                    (jumlah, tgl_masuk, tgl_masuk, buku['id'])
                )
                berhasil.append(f"{buku['judul']} (+{jumlah})")
            except Exception:
                gagal.append(f"Baris {baris_ke}: gagal memproses ISBN {isbn}")

        conn.commit()
        cur.close()
        conn.close()

        pesan = f'{len(berhasil)} buku berhasil diupdate stoknya.'
        if gagal:
            pesan += f' {len(gagal)} baris gagal/dilewati.'
        flash(pesan, 'success' if berhasil else 'danger')

        return render_template('transaksi/import_masuk.html', berhasil=berhasil, gagal=gagal)

    return render_template('transaksi/import_masuk.html')

# ------------------ PROGRESS PENERIMAAN PER PENERBIT ------------------
@app.route('/buku/progress-penerbit')
@login_required
@viewer_blocked
def progress_penerbit():
    sort = request.args.get('sort', 'eks_asc')

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """SELECT 
             COALESCE(NULLIF(penerbit, ''), '(Tanpa Penerbit)') as penerbit,
             COUNT(*) as total_judul,
             COUNT(*) FILTER (WHERE jumlah_rencana > 0 AND stok = jumlah_rencana) as judul_lengkap,
             COUNT(*) FILTER (WHERE jumlah_rencana > 0 AND stok > jumlah_rencana) as judul_lebih,
             COUNT(*) FILTER (WHERE jumlah_rencana > 0 AND stok > 0 AND stok < jumlah_rencana) as judul_sebagian,
             COUNT(*) FILTER (WHERE jumlah_rencana > 0 AND stok = 0) as judul_kosong,
             COALESCE(SUM(stok), 0) as total_diterima,
             COALESCE(SUM(jumlah_rencana), 0) as total_rencana
           FROM buku
           GROUP BY COALESCE(NULLIF(penerbit, ''), '(Tanpa Penerbit)')
           ORDER BY penerbit ASC"""
    )
    data_penerbit = cur.fetchall()
    cur.close()
    conn.close()

    hasil = []
    for p in data_penerbit:
        persen_eks = (p['total_diterima'] / p['total_rencana'] * 100) if p['total_rencana'] > 0 else 0
        hasil.append({
            'penerbit': p['penerbit'],
            'total_judul': p['total_judul'],
            'judul_lengkap': p['judul_lengkap'],
            'judul_lebih': p['judul_lebih'],
            'judul_sebagian': p['judul_sebagian'],
            'judul_kosong': p['judul_kosong'],
            'total_diterima': p['total_diterima'],
            'total_rencana': p['total_rencana'],
            'persen_eks': round(persen_eks, 1)
        })

    sort_map = {
        'penerbit': lambda x: x['penerbit'].lower(),
        'eks_asc': lambda x: x['persen_eks'],
        'eks_desc': lambda x: -x['persen_eks'],
        'lengkap': lambda x: -x['judul_lengkap'],
        'lebih': lambda x: -x['judul_lebih'],
        'sebagian': lambda x: -x['judul_sebagian'],
        'kosong': lambda x: -x['judul_kosong'],
        'total': lambda x: -x['total_judul'],
    }
    hasil.sort(key=sort_map.get(sort, sort_map['eks_asc']))

    return render_template('buku/progress_penerbit.html', daftar=hasil, sort=sort)

# ------------------ ADMIN: KELOLA PENERBIT ------------------
@app.route('/admin/penerbit')
@login_required
@admin_required
def kelola_penerbit():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """SELECT COALESCE(NULLIF(penerbit, ''), '(Tanpa Penerbit)') as penerbit, COUNT(*) as jumlah_buku
           FROM buku
           GROUP BY COALESCE(NULLIF(penerbit, ''), '(Tanpa Penerbit)')
           ORDER BY penerbit ASC"""
    )
    daftar_penerbit = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('admin/penerbit.html', daftar_penerbit=daftar_penerbit)


# ------------------ ADMIN: GABUNG/GANTI NAMA PENERBIT ------------------
@app.route('/admin/penerbit/gabung', methods=['POST'])
@login_required
@admin_required
def gabung_penerbit():
    penerbit_lama = request.form.get('penerbit_lama', '').strip()
    penerbit_baru = request.form.get('penerbit_baru', '').strip()

    if not penerbit_lama or not penerbit_baru:
        flash('Pilih penerbit lama dan isi nama penerbit baru.', 'danger')
        return redirect(url_for('kelola_penerbit'))

    if penerbit_lama == penerbit_baru:
        flash('Nama baru sama dengan nama lama, tidak ada yang diubah.', 'danger')
        return redirect(url_for('kelola_penerbit'))

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute(
            "UPDATE buku SET penerbit = %s WHERE penerbit = %s",
            (penerbit_baru, penerbit_lama)
        )
        jumlah_terupdate = cur.rowcount
        conn.commit()
        print(f"[GABUNG PENERBIT] oleh {session.get('username')} pada {datetime.now()}: '{penerbit_lama}' -> '{penerbit_baru}' ({jumlah_terupdate} buku)")
        catat_aktivitas('Menggabung Penerbit', f'"{penerbit_lama}" digabung ke "{penerbit_baru}" ({jumlah_terupdate} buku terpengaruh)')
        flash(f'{jumlah_terupdate} buku dari "{penerbit_lama}" berhasil diubah jadi "{penerbit_baru}".', 'success')
    except Exception as e:
        conn.rollback()
        flash('Gagal memproses. Coba lagi.', 'danger')
    finally:
        cur.close()
        conn.close()

    return redirect(url_for('kelola_penerbit'))

@app.route('/buku/sync-sheets', methods=['POST'])
@login_required
@viewer_blocked
@admin_required
def buku_sync_sheets():
    sukses, pesan = sync_ke_google_sheets()
    flash(pesan, 'success' if sukses else 'danger')
    return redirect(url_for('buku_list'))   

@app.errorhandler(413)
def file_terlalu_besar(e):
    flash('File terlalu besar. Maksimal ukuran file adalah 5MB.', 'danger')
    return redirect(request.referrer or url_for('dashboard'))

# ------------------ BACKUP MANUAL - EXCEL ------------------
@app.route('/admin/backup/excel')
@login_required
@admin_required
def backup_excel():
    output = buat_backup_excel()
    filename = f"backup-inventory-gudang-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ------------------ BACKUP MANUAL - JSON (untuk restore) ------------------
@app.route('/admin/backup/json')
@login_required
@admin_required
def backup_json():
    output = buat_backup_json()
    filename = f"backup-inventory-gudang-{datetime.now().strftime('%Y%m%d-%H%M')}.json"
    return send_file(
        output,
        mimetype='application/json',
        as_attachment=True,
        download_name=filename
    )


# ------------------ BACKUP OTOMATIS TERJADWAL (dipanggil cron) ------------------
@app.route('/cron/backup-otomatis')
def cron_backup_otomatis():
    secret = request.args.get('secret', '')
    if secret != os.environ.get('CRON_SECRET'):
        return {'error': 'unauthorized'}, 401

    sukses, pesan = kirim_backup_email()
    return {'success': sukses, 'message': pesan}

# ------------------ ADMIN: HALAMAN RESTORE ------------------
@app.route('/admin/restore', methods=['GET', 'POST'])
@login_required
@admin_required
def restore_backup():
    if request.method == 'POST':
        konfirmasi = request.form.get('konfirmasi', '').strip()
        password_konfirmasi = request.form.get('password_konfirmasi', '')
        file = request.files.get('file_backup')

        if konfirmasi != 'RESTORE':
            flash('Ketik "RESTORE" persis (huruf besar semua) untuk konfirmasi.', 'danger')
            return render_template('admin/restore.html')

        conn_cek = get_db_connection()
        cur_cek = conn_cek.cursor()
        cur_cek.execute("SELECT password_hash FROM users WHERE id = %s", (session['user_id'],))
        user_sekarang = cur_cek.fetchone()
        cur_cek.close()
        conn_cek.close()

        if not user_sekarang or not check_password_hash(user_sekarang['password_hash'], password_konfirmasi):
            flash('Password konfirmasi salah.', 'danger')
            return render_template('admin/restore.html')

        if not file or file.filename == '':
            flash('Pilih file backup JSON dulu.', 'danger')
            return render_template('admin/restore.html')

        if not file.filename.endswith('.json'):
            flash('File harus berformat .json (hasil dari fitur Backup JSON).', 'danger')
            return render_template('admin/restore.html')

        try:
            data_backup = json_lib.loads(file.read().decode('utf-8'))
        except Exception:
            flash('File JSON tidak valid atau rusak.', 'danger')
            return render_template('admin/restore.html')

        if 'buku' not in data_backup or 'transaksi' not in data_backup or 'users' not in data_backup:
            flash('Struktur file JSON tidak sesuai format backup aplikasi ini.', 'danger')
            return render_template('admin/restore.html')
        catat_aktivitas('Restore Backup', f'Database dipulihkan dari file "{file.filename}"')
        sukses, pesan = restore_dari_backup(data_backup)
        flash(pesan, 'success' if sukses else 'danger')

        if sukses:
            session.clear()
            flash('Silakan login ulang setelah restore.', 'success')
            return redirect(url_for('login'))

        return render_template('admin/restore.html')

    return render_template('admin/restore.html')

# ------------------ ADMIN: LIHAT ACTIVITY LOG ------------------
@app.route('/admin/activity-log')
@login_required
@admin_required
def activity_log():
    aksi_filter = request.args.get('aksi', '').strip()

    conn = get_db_connection()
    cur = conn.cursor()

    query = "SELECT * FROM activity_log WHERE 1=1"
    params = []

    if aksi_filter:
        query += " AND aksi = %s"
        params.append(aksi_filter)

    query += " ORDER BY created_at DESC LIMIT 500"

    cur.execute(query, tuple(params))
    daftar_log_raw = cur.fetchall()

    daftar_log = []
    for log in daftar_log_raw:
        log_dict = dict(log)
        if log_dict.get('created_at'):
            log_dict['created_at'] = log_dict['created_at'] + timedelta(hours=7)
        daftar_log.append(log_dict)

    cur.execute("SELECT DISTINCT aksi FROM activity_log ORDER BY aksi ASC")
    daftar_aksi = [row['aksi'] for row in cur.fetchall()]

    cur.close()
    conn.close()

    return render_template('admin/activity_log.html', daftar_log=daftar_log, daftar_aksi=daftar_aksi, aksi_filter=aksi_filter)

# ------------------ IMPORT TUJUAN DARI EXCEL ------------------
@app.route('/tujuan/import', methods=['GET', 'POST'])
@login_required
@admin_required
def tujuan_import():
    if request.method == 'POST':
        file = request.files.get('file_excel')

        if not file or file.filename == '':
            flash('Pilih file Excel dulu.', 'danger')
            return render_template('tujuan/import.html')

        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('File harus berformat .xlsx atau .xls', 'danger')
            return render_template('tujuan/import.html')

        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.worksheets[0]
        except Exception:
            flash('Gagal membaca file Excel. Pastikan formatnya benar.', 'danger')
            return render_template('tujuan/import.html')

        header_row = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]

        alias_kolom = {
            'nama': ['nama perpustakaan desa/kelurahan', 'nama perpustakaan', 'nama'],
            'provinsi': ['provinsi'],
            'kabupaten_kota': ['kabupaten/kota', 'kabupaten', 'kota'],
            'kecamatan': ['kecamatan'],
            'desa_kelurahan': ['desa/kelurahan', 'desa', 'kelurahan'],
        }
        kolom_index = {}
        for field, kemungkinan_nama in alias_kolom.items():
            for nama in kemungkinan_nama:
                if nama in header_row:
                    kolom_index[field] = header_row.index(nama)
                    break

        if 'nama' not in kolom_index:
            flash('Kolom "Nama Perpustakaan" tidak ditemukan di baris pertama file.', 'danger')
            return render_template('tujuan/import.html')

        conn = get_db_connection()
        cur = conn.cursor()

        berhasil = 0
        dilewati = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            def ambil(field):
                idx = kolom_index.get(field)
                if idx is None or idx >= len(row) or row[idx] is None:
                    return ''
                return str(row[idx]).strip()

            nama = ambil('nama')
            if not nama:
                continue

            provinsi = ambil('provinsi')
            kabupaten_kota = ambil('kabupaten_kota')
            kecamatan = ambil('kecamatan')
            desa_kelurahan = ambil('desa_kelurahan')

            cur.execute(
                """SELECT id FROM tujuan 
                   WHERE nama = %s AND COALESCE(kecamatan, '') = %s AND COALESCE(desa_kelurahan, '') = %s""",
                (nama, kecamatan, desa_kelurahan)
            )
            if cur.fetchone():
                dilewati += 1
                continue

            cur.execute(
                """INSERT INTO tujuan (nama, provinsi, kabupaten_kota, kecamatan, desa_kelurahan)
                   VALUES (%s, %s, %s, %s, %s)""",
                (nama, ambil('provinsi'), ambil('kabupaten_kota'), ambil('kecamatan'), ambil('desa_kelurahan'))
            )
            berhasil += 1

        conn.commit()
        cur.close()
        conn.close()

        catat_aktivitas('Import Tujuan', f'{berhasil} tujuan baru ditambahkan, {dilewati} dilewati (sudah ada)')
        flash(f'{berhasil} tujuan berhasil diimpor. {dilewati} dilewati (nama sudah ada).', 'success')
        return redirect(url_for('tujuan_list'))

    return render_template('tujuan/import.html')

# ------------------ ADMIN: DAFTAR TUJUAN ------------------
@app.route('/tujuan')
@login_required
@viewer_blocked
def tujuan_list():
    search = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '').strip()
    sort = request.args.get('sort', 'nama')

    conn = get_db_connection()
    cur = conn.cursor()

    query = """
        SELECT t.*,
             COALESCE(SUM(dr.jumlah_rencana), 0) as total_rencana,
             COUNT(DISTINCT dr.buku_id) as total_judul_rencana,
             COALESCE((
                 SELECT SUM(tr.jumlah) FROM transaksi tr
                 WHERE tr.tujuan_id = t.id AND tr.tipe = 'keluar'
             ), 0) as total_terkirim
        FROM tujuan t
        LEFT JOIN distribusi_rencana dr ON dr.tujuan_id = t.id
        WHERE 1=1
    """
    params = []

    if search:
        query += " AND (t.nama ILIKE %s OR t.kecamatan ILIKE %s OR t.kabupaten_kota ILIKE %s)"
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])

    query += " GROUP BY t.id"

    cur.execute(query, tuple(params))
    daftar_tujuan_raw = cur.fetchall()
    cur.close()
    conn.close()

    daftar_tujuan = []
    for t in daftar_tujuan_raw:
        persen = (t['total_terkirim'] / t['total_rencana'] * 100) if t['total_rencana'] > 0 else 0
        if t['total_rencana'] == 0:
            status = 'Belum Ada Rencana'
        elif t['total_terkirim'] == 0:
            status = 'Belum Dikirim'
        elif t['total_terkirim'] < t['total_rencana']:
            status = 'Sebagian'
        else:
            status = 'Lengkap'

        item = dict(t)
        item['persen'] = round(persen, 1)
        item['status'] = status
        daftar_tujuan.append(item)

    if status_filter:
        daftar_tujuan = [t for t in daftar_tujuan if t['status'] == status_filter]

    sort_map = {
        'nama': lambda x: x['nama'].lower(),
        'persen_asc': lambda x: x['persen'],
        'persen_desc': lambda x: -x['persen'],  
        'rencana': lambda x: -x['total_rencana'],
    }
    daftar_tujuan.sort(key=sort_map.get(sort, sort_map['nama']))

    return render_template('tujuan/list.html', daftar_tujuan=daftar_tujuan, search=search, status_filter=status_filter, sort=sort)


# ------------------ ADMIN: TAMBAH TUJUAN ------------------
@app.route('/tujuan/tambah', methods=['GET', 'POST'])
@login_required
@admin_required
def tujuan_tambah():
    if request.method == 'POST':
        nama = request.form.get('nama', '').strip()
        provinsi = request.form.get('provinsi', '').strip()
        kabupaten_kota = request.form.get('kabupaten_kota', '').strip()
        kecamatan = request.form.get('kecamatan', '').strip()
        desa_kelurahan = request.form.get('desa_kelurahan', '').strip()
        alamat = request.form.get('alamat', '').strip()
        catatan = request.form.get('catatan', '').strip()

        if not nama:
            flash('Nama tujuan wajib diisi.', 'danger')
            return render_template('tujuan/form.html')

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("SELECT id FROM tujuan WHERE nama = %s", (nama,))
        if cur.fetchone():
            flash('Tujuan dengan nama ini sudah ada.', 'danger')
            cur.close()
            conn.close()
            return render_template('tujuan/form.html')

        cur.execute(
            """INSERT INTO tujuan (nama, provinsi, kabupaten_kota, kecamatan, desa_kelurahan, alamat, catatan)
               VALUES (%s, %s, %s, %s, %s, %s, %s)""",
            (nama, provinsi, kabupaten_kota, kecamatan, desa_kelurahan, alamat, catatan)
        )
        conn.commit()
        cur.close()
        conn.close()

        catat_aktivitas('Menambah Tujuan', f'Tujuan "{nama}" ditambahkan')
        flash(f'Tujuan "{nama}" berhasil ditambahkan.', 'success')
        return redirect(url_for('tujuan_list'))

    return render_template('tujuan/form.html')


# ------------------ ADMIN: EDIT TUJUAN ------------------
@app.route('/tujuan/edit/<int:tujuan_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def tujuan_edit(tujuan_id):
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == 'POST':
        nama = request.form.get('nama', '').strip()
        provinsi = request.form.get('provinsi', '').strip()
        kabupaten_kota = request.form.get('kabupaten_kota', '').strip()
        kecamatan = request.form.get('kecamatan', '').strip()
        desa_kelurahan = request.form.get('desa_kelurahan', '').strip()
        alamat = request.form.get('alamat', '').strip()
        catatan = request.form.get('catatan', '').strip()

        if not nama:
            flash('Nama tujuan wajib diisi.', 'danger')
            cur.close()
            conn.close()
            return render_template('tujuan/form.html', tujuan=request.form, tujuan_id=tujuan_id)

        cur.execute(
            """UPDATE tujuan SET nama=%s, provinsi=%s, kabupaten_kota=%s, kecamatan=%s, desa_kelurahan=%s, alamat=%s, catatan=%s
               WHERE id=%s""",
            (nama, provinsi, kabupaten_kota, kecamatan, desa_kelurahan, alamat, catatan, tujuan_id)
        )
        conn.commit()
        cur.close()
        conn.close()

        catat_aktivitas('Mengedit Tujuan', f'Tujuan "{nama}" diperbarui')
        flash(f'Tujuan "{nama}" berhasil diupdate.', 'success')
        return redirect(url_for('tujuan_list'))

    cur.execute("SELECT * FROM tujuan WHERE id = %s", (tujuan_id,))
    tujuan = cur.fetchone()
    cur.close()
    conn.close()

    if not tujuan:
        flash('Tujuan tidak ditemukan.', 'danger')
        return redirect(url_for('tujuan_list'))

    return render_template('tujuan/form.html', tujuan=tujuan, tujuan_id=tujuan_id)


# ------------------ ADMIN: HAPUS TUJUAN ------------------
@app.route('/tujuan/hapus/<int:tujuan_id>', methods=['POST'])
@login_required
@admin_required
def tujuan_hapus(tujuan_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT nama FROM tujuan WHERE id = %s", (tujuan_id,))
    tujuan = cur.fetchone()

    if not tujuan:
        flash('Tujuan tidak ditemukan.', 'danger')
    else:
        try:
            cur.execute("DELETE FROM distribusi_rencana WHERE tujuan_id = %s", (tujuan_id,))
            cur.execute("DELETE FROM tujuan WHERE id = %s", (tujuan_id,))
            conn.commit()
            catat_aktivitas('Menghapus Tujuan', f'Tujuan "{tujuan["nama"]}" dihapus')
            flash(f'Tujuan "{tujuan["nama"]}" berhasil dihapus.', 'success')
        except Exception:
            conn.rollback()
            flash('Gagal menghapus tujuan — mungkin masih ada transaksi terkait.', 'danger')

    cur.close()
    conn.close()
    return redirect(url_for('tujuan_list'))

# ------------------ DETAIL TUJUAN + RENCANA DISTRIBUSI ------------------
@app.route('/tujuan/<int:tujuan_id>/detail')
@login_required
@viewer_blocked
def tujuan_detail(tujuan_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM tujuan WHERE id = %s", (tujuan_id,))
    tujuan = cur.fetchone()

    if not tujuan:
        cur.close()
        conn.close()
        flash('Tujuan tidak ditemukan.', 'danger')
        return redirect(url_for('tujuan_list'))

    cur.execute(
        """SELECT dr.*, b.isbn, b.judul, b.penerbit,
             COALESCE((
                 SELECT SUM(t.jumlah) FROM transaksi t
                 WHERE t.tujuan_id = dr.tujuan_id AND t.buku_id = dr.buku_id AND t.tipe = 'keluar'
             ), 0) as jumlah_terkirim
           FROM distribusi_rencana dr
           JOIN buku b ON dr.buku_id = b.id
           WHERE dr.tujuan_id = %s
           ORDER BY b.judul ASC""",
        (tujuan_id,)
    )
    rencana = cur.fetchall()

    cur.close()
    conn.close()

    total_rencana = sum(r['jumlah_rencana'] for r in rencana)
    total_terkirim = sum(r['jumlah_terkirim'] for r in rencana)

    return render_template('tujuan/detail.html', tujuan=tujuan, rencana=rencana,
                            total_rencana=total_rencana, total_terkirim=total_terkirim)


# ------------------ IMPORT RENCANA DISTRIBUSI UNTUK 1 TUJUAN ------------------
@app.route('/tujuan/<int:tujuan_id>/import-rencana', methods=['GET', 'POST'])
@login_required
@admin_required
def tujuan_import_rencana(tujuan_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM tujuan WHERE id = %s", (tujuan_id,))
    tujuan = cur.fetchone()

    if not tujuan:
        cur.close()
        conn.close()
        flash('Tujuan tidak ditemukan.', 'danger')
        return redirect(url_for('tujuan_list'))

    if request.method == 'POST':
        file = request.files.get('file_excel')

        if not file or file.filename == '':
            flash('Pilih file Excel dulu.', 'danger')
            cur.close()
            conn.close()
            return render_template('tujuan/import_rencana.html', tujuan=tujuan)

        try:
            semua_rows = baca_excel_universal(file)
        except Exception as e:
            flash(f'Gagal membaca file Excel: {str(e)}', 'danger')
            cur.close()
            conn.close()
            return render_template('tujuan/import_rencana.html', tujuan=tujuan)

        # cari baris header (file "Laporan Per Faktur" ada judul di baris atas sebelum tabel)
        header_row_idx = None
        kolom_index = {}
        alias_kolom = {
            'isbn': ['isbn'],
            'eksemplar': ['eksemplar', 'jumlah eksemplar', 'jumlah'],
        }

        for i, row in enumerate(semua_rows[:15]):
            row_values = [str(cell).strip().lower() if cell is not None else '' for cell in row]
            if 'isbn' in row_values:
                header_row_idx = i
                for field, kemungkinan_nama in alias_kolom.items():
                    for nama in kemungkinan_nama:
                        if nama in row_values:
                            kolom_index[field] = row_values.index(nama)
                            break
                break

        if header_row_idx is None or 'isbn' not in kolom_index:
            flash('Kolom ISBN tidak ditemukan di file.', 'danger')
            cur.close()
            conn.close()
            return render_template('tujuan/import_rencana.html', tujuan=tujuan)

        berhasil = 0
        tidak_ketemu = []
        agregasi_per_buku = {}

        for row in semua_rows[header_row_idx + 1:]:
            def ambil(field, default=''):
                idx = kolom_index.get(field)
                if idx is None or idx >= len(row) or row[idx] is None:
                    return default
                return row[idx]

            isbn = str(ambil('isbn')).strip()
            if not isbn or isbn.lower() == 'none':
                continue

            try:
                eksemplar = int(float(ambil('eksemplar', 1) or 1))
            except (ValueError, TypeError):
                eksemplar = 1

            cur.execute("SELECT id FROM buku WHERE isbn = %s", (isbn,))
            buku = cur.fetchone()

            if not buku:
                tidak_ketemu.append(isbn)
                continue

            agregasi_per_buku[buku['id']] = agregasi_per_buku.get(buku['id'], 0) + eksemplar

        for buku_id, total_eksemplar in agregasi_per_buku.items():
            cur.execute(
                """INSERT INTO distribusi_rencana (tujuan_id, buku_id, jumlah_rencana)
                   VALUES (%s, %s, %s)
                   ON CONFLICT (tujuan_id, buku_id)
                   DO UPDATE SET jumlah_rencana = EXCLUDED.jumlah_rencana""",
                (tujuan_id, buku_id, total_eksemplar)
            )
            berhasil += 1

        conn.commit()
        cur.close()
        conn.close()

        catat_aktivitas('Import Rencana Distribusi', f'{berhasil} judul untuk tujuan "{tujuan["nama"]}"')
        pesan = f'{berhasil} rencana distribusi berhasil diimpor untuk "{tujuan["nama"]}".'
        if tidak_ketemu:
            pesan += f' {len(tidak_ketemu)} ISBN tidak ditemukan di master data buku.'
        flash(pesan, 'success')
        return redirect(url_for('tujuan_detail', tujuan_id=tujuan_id))

    cur.close()
    conn.close()
    return render_template('tujuan/import_rencana.html', tujuan=tujuan)

# ------------------ IMPORT RENCANA DISTRIBUSI MASSAL (BANYAK FILE SEKALIGUS) ------------------
@app.route('/tujuan/import-rencana-massal', methods=['GET', 'POST'])
@login_required
@admin_required
def tujuan_import_rencana_massal():
    if request.method == 'POST':
        files = request.files.getlist('files')

        if not files or files[0].filename == '':
            flash('Pilih minimal satu file.', 'danger')
            return render_template('tujuan/import_rencana_massal.html')

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("SELECT id, nama FROM tujuan")
        semua_tujuan = cur.fetchall()
        peta_tujuan = {}
        for t in semua_tujuan:
            key = re.sub(r'\s+', ' ', t['nama'].strip().lower())
            peta_tujuan.setdefault(key, []).append(t['id'])
        peta_nama_by_id = {t['id']: t['nama'] for t in semua_tujuan}

        cur.execute("SELECT id, isbn FROM buku")
        peta_buku = {b['isbn'].strip(): b['id'] for b in cur.fetchall()}

        file_tujuan_tidak_ketemu = []
        file_nama_bentrok = []
        file_gagal_baca = []
        ringkasan_per_file = []  # (nama_file, nama_tujuan, jumlah_judul, total_eksemplar, isbn_dobel, isbn_tidak_ketemu)
        gabungan_staging = {}  # (tujuan_id, buku_id) -> {'jumlah': int, 'files': set()}
        tujuan_id_ke_file = {}

        for f in files:
            nama_file = f.filename
            if not nama_file or not nama_file.lower().endswith(('.xls', '.xlsx')):
                continue

            nama_tanpa_ext = re.sub(r'\.(xls|xlsx)$', '', nama_file, flags=re.IGNORECASE)
            bagian_nama = re.sub(r'^\d+[\.\s_\-]+', '', nama_tanpa_ext)
            nama_tujuan_dicari = re.sub(r'[_\s]+', ' ', bagian_nama).strip().lower()

            daftar_id_cocok = peta_tujuan.get(nama_tujuan_dicari, [])

            if len(daftar_id_cocok) == 0:
                file_tujuan_tidak_ketemu.append(nama_file)
                continue
            if len(daftar_id_cocok) > 1:
                file_nama_bentrok.append(f'{nama_file} → cocok dengan {len(daftar_id_cocok)} tujuan sekaligus')
                continue

            tujuan_id = daftar_id_cocok[0]
            tujuan_id_ke_file.setdefault(tujuan_id, []).append(nama_file)

            try:
                semua_rows = baca_excel_universal(f)
            except Exception:
                file_gagal_baca.append(nama_file)
                continue

            header_row_idx = None
            kolom_index = {}
            alias_kolom = {'isbn': ['isbn'], 'eksemplar': ['eksemplar', 'jumlah eksemplar', 'jumlah']}

            for i, row in enumerate(semua_rows[:15]):
                row_values = [str(cell).strip().lower() if cell is not None else '' for cell in row]
                if 'isbn' in row_values:
                    header_row_idx = i
                    for field, kemungkinan_nama in alias_kolom.items():
                        for nama in kemungkinan_nama:
                            if nama in row_values:
                                kolom_index[field] = row_values.index(nama)
                                break
                    break

            if header_row_idx is None or 'isbn' not in kolom_index:
                file_gagal_baca.append(nama_file)
                continue

            agregasi_file_ini = {}
            isbn_muncul = {}
            isbn_tdk_ketemu_file = 0

            for row in semua_rows[header_row_idx + 1:]:
                idx_isbn = kolom_index.get('isbn')
                idx_eks = kolom_index.get('eksemplar')

                if idx_isbn is None or idx_isbn >= len(row) or row[idx_isbn] is None:
                    continue

                isbn = str(row[idx_isbn]).strip()
                if not isbn:
                    continue

                try:
                    eksemplar = int(float(row[idx_eks])) if idx_eks is not None and idx_eks < len(row) and row[idx_eks] else 1
                except (ValueError, TypeError):
                    eksemplar = 1

                isbn_muncul[isbn] = isbn_muncul.get(isbn, 0) + 1

                buku_id = peta_buku.get(isbn)
                if not buku_id:
                    isbn_tdk_ketemu_file += 1
                    continue

                agregasi_file_ini[buku_id] = agregasi_file_ini.get(buku_id, 0) + eksemplar

            isbn_dobel = [isbn for isbn, jumlah in isbn_muncul.items() if jumlah > 1]

            for buku_id, jumlah in agregasi_file_ini.items():
                key = (tujuan_id, buku_id)
                if key not in gabungan_staging:
                    gabungan_staging[key] = {'jumlah': 0, 'files': set()}
                gabungan_staging[key]['jumlah'] += jumlah
                gabungan_staging[key]['files'].add(nama_file)

            ringkasan_per_file.append((
                nama_file, peta_nama_by_id.get(tujuan_id, '?'),
                len(agregasi_file_ini), sum(agregasi_file_ini.values()),
                isbn_dobel, isbn_tdk_ketemu_file
            ))

        # simpan hasil parsing ke staging (belum masuk ke distribusi_rencana)
        batch_id = secrets_lib.token_hex(8)
        for (tujuan_id, buku_id), info in gabungan_staging.items():
            cur.execute(
                """INSERT INTO import_staging (batch_id, tujuan_id, buku_id, jumlah_rencana, nama_file)
                   VALUES (%s, %s, %s, %s, %s)""",
                (batch_id, tujuan_id, buku_id, info['jumlah'], ', '.join(info['files']))
            )
        conn.commit()
        cur.close()
        conn.close()

        tujuan_dobel_file = {tid: fls for tid, fls in tujuan_id_ke_file.items() if len(fls) > 1}
        detail_tujuan_dobel = [(peta_nama_by_id.get(tid, '?'), fls) for tid, fls in tujuan_dobel_file.items()]

        total_tujuan_terpengaruh = len(tujuan_id_ke_file)
        total_eksemplar_preview = sum(info['jumlah'] for info in gabungan_staging.values())

        return render_template(
            'tujuan/import_rencana_preview.html',
            batch_id=batch_id,
            ringkasan_per_file=ringkasan_per_file,
            file_tujuan_tidak_ketemu=file_tujuan_tidak_ketemu,
            file_nama_bentrok=file_nama_bentrok,
            file_gagal_baca=file_gagal_baca,
            detail_tujuan_dobel=detail_tujuan_dobel,
            total_tujuan_terpengaruh=total_tujuan_terpengaruh,
            total_eksemplar_preview=total_eksemplar_preview
        )

    return render_template('tujuan/import_rencana_massal.html')


# ------------------ KONFIRMASI: SIMPAN HASIL PRATINJAU KE DATABASE ------------------
@app.route('/tujuan/import-rencana-massal/konfirmasi', methods=['POST'])
@login_required
@admin_required
def tujuan_import_rencana_konfirmasi():
    batch_id = request.form.get('batch_id', '').strip()
    if not batch_id:
        flash('Batch tidak valid.', 'danger')
        return redirect(url_for('tujuan_list'))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM import_staging WHERE batch_id = %s", (batch_id,))
    staging_rows = cur.fetchall()

    if not staging_rows:
        cur.close()
        conn.close()
        flash('Data pratinjau sudah tidak ada (mungkin sudah dikonfirmasi/dibatalkan sebelumnya).', 'danger')
        return redirect(url_for('tujuan_list'))

    for row in staging_rows:
        cur.execute(
            """INSERT INTO distribusi_rencana (tujuan_id, buku_id, jumlah_rencana)
               VALUES (%s, %s, %s)
               ON CONFLICT (tujuan_id, buku_id)
               DO UPDATE SET jumlah_rencana = EXCLUDED.jumlah_rencana""",
            (row['tujuan_id'], row['buku_id'], row['jumlah_rencana'])
        )

    cur.execute("DELETE FROM import_staging WHERE batch_id = %s", (batch_id,))
    conn.commit()
    cur.close()
    conn.close()

    catat_aktivitas('Import Rencana Massal (Konfirmasi)', f'{len(staging_rows)} baris rencana dikonfirmasi dan disimpan')
    flash(f'{len(staging_rows)} rencana distribusi berhasil disimpan.', 'success')
    return redirect(url_for('tujuan_list'))


# ------------------ BATALKAN PRATINJAU ------------------
@app.route('/tujuan/import-rencana-massal/batal', methods=['POST'])
@login_required
@admin_required
def tujuan_import_rencana_batal():
    batch_id = request.form.get('batch_id', '').strip()
    if batch_id:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM import_staging WHERE batch_id = %s", (batch_id,))
        conn.commit()
        cur.close()
        conn.close()

    flash('Import dibatalkan, tidak ada perubahan data.', 'success')
    return redirect(url_for('tujuan_import_rencana_massal'))

# ------------------ IMPORT MASSAL PENGIRIMAN (BUKU KELUAR PER TUJUAN) ------------------
@app.route('/tujuan/import-pengiriman-massal', methods=['GET', 'POST'])
@login_required
@admin_required
def tujuan_import_pengiriman_massal():
    if request.method == 'POST':
        files = request.files.getlist('files')
        tanggal_kirim = request.form.get('tanggal_kirim', '').strip() or datetime.now().date().isoformat()

        if not files or files[0].filename == '':
            flash('Pilih minimal satu file.', 'danger')
            return render_template('tujuan/import_pengiriman_massal.html', today=datetime.now().date().isoformat())

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("SELECT id, nama FROM tujuan")
        semua_tujuan = cur.fetchall()
        peta_tujuan = {}
        for t in semua_tujuan:
            key = re.sub(r'\s+', ' ', t['nama'].strip().lower())
            peta_tujuan.setdefault(key, []).append(t['id'])

        cur.execute("SELECT id, isbn, stok FROM buku")
        peta_buku = {b['isbn'].strip(): {'id': b['id'], 'stok': b['stok']} for b in cur.fetchall()}

        file_berhasil = 0
        file_tujuan_tidak_ketemu = []
        file_nama_bentrok = []
        file_gagal_baca = []
        file_ada_duplikat = []
        file_stok_kurang = []  # (nama_file, [(isbn, diminta, tersedia), ...])
        total_baris_masuk = 0
        total_isbn_tidak_ketemu = 0

        for f in files:
            nama_file = f.filename
            if not nama_file or not nama_file.lower().endswith(('.xls', '.xlsx')):
                continue

            nama_tanpa_ext = re.sub(r'\.(xls|xlsx)$', '', nama_file, flags=re.IGNORECASE)
            bagian_nama = re.sub(r'^\d+[\.\s_\-]+', '', nama_tanpa_ext)
            nama_tujuan_dicari = re.sub(r'[_\s]+', ' ', bagian_nama).strip().lower()

            daftar_id_cocok = peta_tujuan.get(nama_tujuan_dicari, [])

            if len(daftar_id_cocok) == 0:
                file_tujuan_tidak_ketemu.append(nama_file)
                continue
            if len(daftar_id_cocok) > 1:
                file_nama_bentrok.append(f'{nama_file} → cocok dengan {len(daftar_id_cocok)} tujuan sekaligus')
                continue

            tujuan_id = daftar_id_cocok[0]

            try:
                semua_rows = baca_excel_universal(f)
            except Exception:
                file_gagal_baca.append(nama_file)
                continue

            header_row_idx = None
            kolom_index = {}
            alias_kolom = {'isbn': ['isbn'], 'eksemplar': ['eksemplar', 'jumlah eksemplar', 'jumlah']}

            for i, row in enumerate(semua_rows[:15]):
                row_values = [str(cell).strip().lower() if cell is not None else '' for cell in row]
                if 'isbn' in row_values:
                    header_row_idx = i
                    for field, kemungkinan_nama in alias_kolom.items():
                        for nama in kemungkinan_nama:
                            if nama in row_values:
                                kolom_index[field] = row_values.index(nama)
                                break
                    break

            if header_row_idx is None or 'isbn' not in kolom_index:
                file_gagal_baca.append(nama_file)
                continue

            agregasi_per_buku = {}
            isbn_muncul = {}

            for row in semua_rows[header_row_idx + 1:]:
                idx_isbn = kolom_index.get('isbn')
                idx_eks = kolom_index.get('eksemplar')

                if idx_isbn is None or idx_isbn >= len(row) or row[idx_isbn] is None:
                    continue

                isbn = str(row[idx_isbn]).strip()
                if not isbn:
                    continue

                try:
                    eksemplar = int(float(row[idx_eks])) if idx_eks is not None and idx_eks < len(row) and row[idx_eks] else 1
                except (ValueError, TypeError):
                    eksemplar = 1

                isbn_muncul[isbn] = isbn_muncul.get(isbn, 0) + 1

                info_buku = peta_buku.get(isbn)
                if not info_buku:
                    total_isbn_tidak_ketemu += 1
                    continue

                agregasi_per_buku[info_buku['id']] = agregasi_per_buku.get(info_buku['id'], 0) + eksemplar

            isbn_dobel = [isbn for isbn, jumlah in isbn_muncul.items() if jumlah > 1]
            if isbn_dobel:
                file_ada_duplikat.append((nama_file, isbn_dobel))

            # cek stok cukup sebelum proses transaksi
            stok_kurang_file = []
            for buku_id, jumlah_diminta in agregasi_per_buku.items():
                cur.execute("SELECT isbn, stok FROM buku WHERE id = %s", (buku_id,))
                b = cur.fetchone()
                if b and b['stok'] < jumlah_diminta:
                    stok_kurang_file.append((b['isbn'], jumlah_diminta, b['stok']))

            if stok_kurang_file:
                file_stok_kurang.append((nama_file, stok_kurang_file))
                continue  # skip seluruh file ini, jangan proses sebagian

            for buku_id, jumlah in agregasi_per_buku.items():
                cur.execute(
                    """INSERT INTO transaksi (buku_id, tipe, jumlah, user_id, tanggal, tujuan_id, keterangan)
                       VALUES (%s, 'keluar', %s, %s, %s, %s, %s)""",
                    (buku_id, jumlah, session['user_id'], tanggal_kirim, tujuan_id, f'Import massal: {nama_file}')
                )
                cur.execute(
                    "UPDATE buku SET stok = stok - %s, updated_at = NOW() WHERE id = %s",
                    (jumlah, buku_id)
                )
                total_baris_masuk += 1

            file_berhasil += 1

        conn.commit()
        cur.close()
        conn.close()

        catat_aktivitas('Import Pengiriman Massal', f'{file_berhasil} file diproses, {total_baris_masuk} baris pengiriman dicatat')

        pesan = f'{file_berhasil} file berhasil diproses ({total_baris_masuk} baris pengiriman dicatat).'
        if file_tujuan_tidak_ketemu:
            pesan += f' {len(file_tujuan_tidak_ketemu)} file nama tujuannya tidak ditemukan.'
        if file_nama_bentrok:
            pesan += f' {len(file_nama_bentrok)} file nama tujuannya bentrok.'
        if file_gagal_baca:
            pesan += f' {len(file_gagal_baca)} file gagal dibaca/format salah.'
        if total_isbn_tidak_ketemu:
            pesan += f' {total_isbn_tidak_ketemu} baris ISBN tidak ditemukan di Data Buku.'
        if file_ada_duplikat:
            pesan += f' {len(file_ada_duplikat)} file punya ISBN dobel (sudah dijumlahkan).'
        if file_stok_kurang:
            pesan += f' {len(file_stok_kurang)} file DILEWATI karena stok tidak cukup — lihat detail di bawah.'

        return render_template(
            'tujuan/import_pengiriman_massal.html',
            hasil=True, pesan=pesan,
            file_tujuan_tidak_ketemu=file_tujuan_tidak_ketemu,
            file_nama_bentrok=file_nama_bentrok,
            file_gagal_baca=file_gagal_baca,
            file_ada_duplikat=file_ada_duplikat,
            file_stok_kurang=file_stok_kurang
        )

    return render_template('tujuan/import_pengiriman_massal.html', today=datetime.now().date().isoformat())

# ------------------ SCAN CEPAT: CATAT PENGIRIMAN PER SCAN ------------------
@app.route('/tujuan/<int:tujuan_id>/scan', methods=['POST'])
@login_required
@viewer_blocked
def tujuan_scan_kirim(tujuan_id):
    isbn = request.json.get('isbn', '').strip() if request.is_json else request.form.get('isbn', '').strip()

    if not isbn:
        return {'success': False, 'message': 'ISBN kosong.'}, 400

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT id, judul, stok FROM buku WHERE isbn = %s", (isbn,))
    buku = cur.fetchone()

    if not buku:
        cur.close()
        conn.close()
        return {'success': False, 'message': f'ISBN {isbn} tidak ditemukan di Data Buku.'}

    if buku['stok'] <= 0:
        cur.close()
        conn.close()
        return {'success': False, 'message': f'Stok "{buku["judul"]}" sudah habis (0).'}

    cur.execute(
        "SELECT jumlah_rencana FROM distribusi_rencana WHERE tujuan_id = %s AND buku_id = %s",
        (tujuan_id, buku['id'])
    )
    cek_rencana = cur.fetchone()

    if not cek_rencana or cek_rencana['jumlah_rencana'] <= 0:
        cur.close()
        conn.close()
        return {'success': False, 'message': f'❌ "{buku["judul"]}" TIDAK ADA di rencana distribusi tujuan ini. Ditolak.'}

    try:
        # cek apakah sudah ada baris transaksi scan cepat untuk buku+tujuan ini hari ini, kalau ada tambah jumlahnya
        cur.execute(
            """SELECT id, jumlah FROM transaksi 
               WHERE tujuan_id = %s AND buku_id = %s AND tipe = 'keluar' 
                     AND tanggal = CURRENT_DATE AND keterangan = 'Scan Cepat Tujuan'""",
            (tujuan_id, buku['id'])
        )
        existing = cur.fetchone()

        if existing:
            cur.execute("UPDATE transaksi SET jumlah = jumlah + 1 WHERE id = %s", (existing['id'],))
        else:
            cur.execute(
                """INSERT INTO transaksi (buku_id, tipe, jumlah, user_id, tanggal, tujuan_id, keterangan)
                   VALUES (%s, 'keluar', 1, %s, CURRENT_DATE, %s, 'Scan Cepat Tujuan')""",
                (buku['id'], session['user_id'], tujuan_id)
            )

        cur.execute("UPDATE buku SET stok = stok - 1, updated_at = NOW() WHERE id = %s", (buku['id'],))

        jumlah_rencana = cek_rencana['jumlah_rencana']

        cur.execute(
            "SELECT COALESCE(SUM(jumlah), 0) as total FROM transaksi WHERE tujuan_id = %s AND buku_id = %s AND tipe = 'keluar'",
            (tujuan_id, buku['id'])
        )
        jumlah_terkirim = cur.fetchone()['total']

        conn.commit()

        return {
            'success': True,
            'buku_id': buku['id'],
            'judul': buku['judul'],
            'isbn': isbn,
            'jumlah_terkirim': jumlah_terkirim,
            'jumlah_rencana': jumlah_rencana,
            'lebih_dari_rencana': jumlah_terkirim > jumlah_rencana
        }
    except Exception as e:
        conn.rollback()
        return {'success': False, 'message': f'Gagal memproses: {str(e)}'}
    finally:
        cur.close()
        conn.close()

# ------------------ EXPORT LAPORAN 1 TUJUAN - EXCEL ------------------
@app.route('/tujuan/<int:tujuan_id>/export/excel')
@login_required
@viewer_blocked
def tujuan_export_excel(tujuan_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM tujuan WHERE id = %s", (tujuan_id,))
    tujuan = cur.fetchone()

    if not tujuan:
        cur.close()
        conn.close()
        flash('Tujuan tidak ditemukan.', 'danger')
        return redirect(url_for('tujuan_list'))

    cur.execute(
        """SELECT dr.*, b.isbn, b.judul, b.penerbit,
             COALESCE((
                 SELECT SUM(t.jumlah) FROM transaksi t
                 WHERE t.tujuan_id = dr.tujuan_id AND t.buku_id = dr.buku_id AND t.tipe = 'keluar'
             ), 0) as jumlah_terkirim
           FROM distribusi_rencana dr
           JOIN buku b ON dr.buku_id = b.id
           WHERE dr.tujuan_id = %s
           ORDER BY b.judul ASC""",
        (tujuan_id,)
    )
    rencana = cur.fetchall()
    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Surat Jalan"

    ws.merge_cells('A1:F1')
    ws['A1'] = f'SURAT JALAN / CHECKLIST PENGIRIMAN BUKU'
    ws['A1'].font = Font(bold=True, size=14)

    ws.merge_cells('A2:F2')
    ws['A2'] = f'Tujuan: {tujuan["nama"]}'
    ws['A2'].font = Font(bold=True)

    ws.merge_cells('A3:F3')
    ws['A3'] = f'{tujuan["desa_kelurahan"] or "-"}, {tujuan["kecamatan"] or "-"}, {tujuan["kabupaten_kota"] or "-"}, {tujuan["provinsi"] or "-"}'

    ws['A4'] = f'Dicetak: {datetime.now().strftime("%d-%m-%Y %H:%M")}'

    headers = ['No', 'ISBN', 'Judul', 'Penerbit', 'Rencana', 'Terkirim', 'Status', 'Ceklis Fisik']
    ws.append([])
    ws.append(headers)
    header_row_num = ws.max_row
    for cell in ws[header_row_num]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')

    for i, r in enumerate(rencana, start=1):
        if r['jumlah_terkirim'] >= r['jumlah_rencana']:
            status = 'Lengkap'
        elif r['jumlah_terkirim'] > 0:
            status = 'Sebagian'
        else:
            status = 'Belum'
        ws.append([i, r['isbn'], r['judul'], r['penerbit'] or '-', r['jumlah_rencana'], r['jumlah_terkirim'], status, ''])

    lebar_kolom = {'A': 5, 'B': 16, 'C': 45, 'D': 25, 'E': 10, 'F': 10, 'G': 10, 'H': 14}
    for kolom, lebar in lebar_kolom.items():
        ws.column_dimensions[kolom].width = lebar

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nama_file_aman = re.sub(r'[^\w\-]', '_', tujuan['nama'])
    filename = f"surat-jalan-{nama_file_aman}-{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ------------------ EXPORT LAPORAN 1 TUJUAN - PDF ------------------
@app.route('/tujuan/<int:tujuan_id>/export/pdf')
@login_required
@viewer_blocked
def tujuan_export_pdf(tujuan_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM tujuan WHERE id = %s", (tujuan_id,))
    tujuan = cur.fetchone()

    if not tujuan:
        cur.close()
        conn.close()
        flash('Tujuan tidak ditemukan.', 'danger')
        return redirect(url_for('tujuan_list'))

    cur.execute(
        """SELECT dr.*, b.isbn, b.judul, b.penerbit,
             COALESCE((
                 SELECT SUM(t.jumlah) FROM transaksi t
                 WHERE t.tujuan_id = dr.tujuan_id AND t.buku_id = dr.buku_id AND t.tipe = 'keluar'
             ), 0) as jumlah_terkirim
           FROM distribusi_rencana dr
           JOIN buku b ON dr.buku_id = b.id
           WHERE dr.tujuan_id = %s
           ORDER BY b.judul ASC""",
        (tujuan_id,)
    )
    rencana = cur.fetchall()
    cur.close()
    conn.close()

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=landscape(A4),
        topMargin=1*cm, bottomMargin=1*cm, leftMargin=1*cm, rightMargin=1*cm
    )
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Surat Jalan / Checklist Pengiriman Buku", styles['Title']))
    elements.append(Paragraph(f"Tujuan: {tujuan['nama']}", styles['Heading3']))
    elements.append(Paragraph(
        f"{tujuan['desa_kelurahan'] or '-'}, {tujuan['kecamatan'] or '-'}, {tujuan['kabupaten_kota'] or '-'}, {tujuan['provinsi'] or '-'}",
        styles['Normal']
    ))
    elements.append(Paragraph(f"Dicetak: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    style_sel = ParagraphStyle('sel', fontSize=8, leading=10, fontName='Helvetica')
    style_header = ParagraphStyle('header', fontSize=8.5, leading=10, fontName='Helvetica-Bold', textColor=colors.white)

    data = [[
        Paragraph('No', style_header), Paragraph('ISBN', style_header), Paragraph('Judul', style_header),
        Paragraph('Penerbit', style_header), Paragraph('Rencana', style_header), Paragraph('Terkirim', style_header),
        Paragraph('Status', style_header), Paragraph('Ceklis', style_header)
    ]]
    for i, r in enumerate(rencana, start=1):
        if r['jumlah_terkirim'] >= r['jumlah_rencana']:
            status = 'Lengkap'
        elif r['jumlah_terkirim'] > 0:
            status = 'Sebagian'
        else:
            status = 'Belum'
        data.append([
            Paragraph(str(i), style_sel), Paragraph(r['isbn'], style_sel), Paragraph(r['judul'], style_sel),
            Paragraph(r['penerbit'] or '-', style_sel), Paragraph(str(r['jumlah_rencana']), style_sel),
            Paragraph(str(r['jumlah_terkirim']), style_sel), Paragraph(status, style_sel), Paragraph('☐', style_sel)
        ])

    col_widths = [1.0*cm, 2.5*cm, 8.0*cm, 5.0*cm, 2.0*cm, 2.0*cm, 2.0*cm, 1.5*cm]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D6EFD')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)

    doc.build(elements)
    output.seek(0)

    nama_file_aman = re.sub(r'[^\w\-]', '_', tujuan['nama'])
    filename = f"surat-jalan-{nama_file_aman}-{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(output, mimetype='application/pdf', as_attachment=True, download_name=filename)

# ------------------ EXPORT AUDIT LENGKAP - SEMUA RENCANA DISTRIBUSI ------------------
@app.route('/tujuan/audit-export')
@login_required
@admin_required
def tujuan_audit_export():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """SELECT t.nama as nama_tujuan, t.kecamatan, t.kabupaten_kota,
                  b.isbn, b.judul, b.penerbit, dr.jumlah_rencana,
                  COALESCE(SUM(tr.jumlah), 0) as jumlah_terkirim
           FROM distribusi_rencana dr
           JOIN tujuan t ON dr.tujuan_id = t.id
           JOIN buku b ON dr.buku_id = b.id
           LEFT JOIN transaksi tr ON tr.tujuan_id = dr.tujuan_id AND tr.buku_id = dr.buku_id AND tr.tipe = 'keluar'
           GROUP BY t.nama, t.kecamatan, t.kabupaten_kota, b.isbn, b.judul, b.penerbit, dr.jumlah_rencana
           ORDER BY t.nama ASC, b.judul ASC"""
    )
    data = cur.fetchall()
    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Rencana Distribusi"

    headers = ['Tujuan', 'Kecamatan', 'Kabupaten/Kota', 'ISBN', 'Judul', 'Penerbit', 'Rencana', 'Terkirim']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')

    for d in data:
        ws.append([
            d['nama_tujuan'], d['kecamatan'] or '-', d['kabupaten_kota'] or '-',
            d['isbn'], d['judul'], d['penerbit'] or '-', d['jumlah_rencana'], d['jumlah_terkirim']
        ])

    lebar_kolom = {'A': 30, 'B': 20, 'C': 20, 'D': 16, 'E': 40, 'F': 25, 'G': 10, 'H': 10}
    for kolom, lebar in lebar_kolom.items():
        ws.column_dimensions[kolom].width = lebar

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"audit-rencana-distribusi-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    app.run(debug=True)